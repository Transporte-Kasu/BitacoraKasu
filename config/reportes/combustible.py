"""
Reporte automático de combustible.
Email con resumen ejecutivo (top 10, estadísticas) + Excel adjunto.
"""
import io
from datetime import timedelta
from django.utils import timezone
from django.core.mail import EmailMessage
from django.conf import settings
from django.db.models import Sum, Count, Avg, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _get_periodo(periodicidad):
    ahora = timezone.localtime(timezone.now())
    if periodicidad == 'diario':
        fecha_inicio = ahora.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        fecha_fin = ahora
        label = f"Día {(ahora - timedelta(days=1)).strftime('%d/%m/%Y')}"
    elif periodicidad == 'semanal':
        fecha_inicio = ahora - timedelta(days=7)
        fecha_fin = ahora
        label = f"Semana {fecha_inicio.strftime('%d/%m')} — {ahora.strftime('%d/%m/%Y')}"
    else:  # mensual: mes anterior completo
        if ahora.month == 1:
            mes_anterior = ahora.replace(year=ahora.year - 1, month=12, day=1)
        else:
            mes_anterior = ahora.replace(month=ahora.month - 1, day=1)
        fecha_inicio = mes_anterior.replace(hour=0, minute=0, second=0, microsecond=0)
        fecha_fin = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
        label = fecha_inicio.strftime('%B %Y').capitalize()
    return fecha_inicio, fecha_fin, label


def _encabezado_style():
    fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    font = Font(color='FFFFFF', bold=True, size=10)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return fill, font, alignment


def _aplicar_encabezado(ws, headers, row=1):
    fill, font, alignment = _encabezado_style()
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = alignment
    ws.row_dimensions[row].height = 28


def generar_excel_combustible(fecha_inicio, fecha_fin):
    """
    Genera Excel con 2 hojas:
    - Cargas del período (todos los registros COMPLETADO)
    - Top unidades
    """
    from modulos.combustible.models import CargaCombustible
    from modulos.unidades.models import Unidad

    wb = openpyxl.Workbook()

    # ── Hoja 1: Cargas ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Cargas Combustible'

    headers1 = [
        'ID', 'Fecha Inicio', 'Fecha Fin', 'Unidad (Eco.)', 'Placa',
        'Tipo Unidad', 'Despachador', 'Litros', 'Kilometraje',
        'Nivel Inicial', 'Estado Candado', 'Tiempo Carga (min)', 'Estado'
    ]
    _aplicar_encabezado(ws1, headers1)

    cargas = CargaCombustible.objects.select_related(
        'unidad', 'despachador'
    ).filter(
        fecha_hora_inicio__range=(fecha_inicio, fecha_fin),
        estado='COMPLETADO'
    ).order_by('-fecha_hora_inicio')

    fill_alt = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
    for idx, c in enumerate(cargas, start=2):
        data = [
            c.id,
            timezone.localtime(c.fecha_hora_inicio).strftime('%d/%m/%Y %H:%M'),
            timezone.localtime(c.fecha_hora_fin).strftime('%d/%m/%Y %H:%M') if c.fecha_hora_fin else '',
            c.unidad.numero_economico,
            c.unidad.placa,
            c.unidad.get_tipo_display(),
            c.despachador.nombre,
            float(c.cantidad_litros),
            c.kilometraje_actual,
            c.get_nivel_combustible_inicial_display(),
            c.get_estado_candado_anterior_display(),
            c.tiempo_carga_minutos or '',
            c.get_estado_display(),
        ]
        ws1.append(data)
        if idx % 2 == 0:
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=idx, column=col).fill = fill_alt

    anchos1 = [8, 18, 18, 14, 12, 14, 22, 10, 14, 16, 18, 16, 12]
    for i, w in enumerate(anchos1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Top Unidades ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Top Unidades')
    headers2 = ['Posición', 'Número Económico', 'Placa', 'Tipo', 'Cargas', 'Total Litros', 'Promedio por Carga (L)']
    _aplicar_encabezado(ws2, headers2)

    top_unidades = (
        CargaCombustible.objects
        .filter(fecha_hora_inicio__range=(fecha_inicio, fecha_fin), estado='COMPLETADO')
        .values('unidad__numero_economico', 'unidad__placa', 'unidad__tipo')
        .annotate(total_litros=Sum('cantidad_litros'), num_cargas=Count('id'), promedio=Avg('cantidad_litros'))
        .order_by('-total_litros')
    )

    for pos, u in enumerate(top_unidades, start=1):
        row_data = [
            pos,
            u['unidad__numero_economico'],
            u['unidad__placa'],
            u['unidad__tipo'],
            u['num_cargas'],
            float(u['total_litros']),
            round(float(u['promedio']), 2),
        ]
        ws2.append(row_data)
        if pos % 2 == 0:
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=pos + 1, column=col).fill = fill_alt

    anchos2 = [10, 18, 14, 14, 10, 16, 22]
    for i, w in enumerate(anchos2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_cuerpo_reporte(fecha_inicio, fecha_fin, label):
    """Genera el cuerpo del correo con estadísticas y top 10."""
    from modulos.combustible.models import CargaCombustible
    from modulos.unidades.models import Unidad

    cargas = CargaCombustible.objects.select_related('unidad').filter(
        fecha_hora_inicio__range=(fecha_inicio, fecha_fin),
        estado='COMPLETADO'
    )

    # ── Estadísticas generales ────────────────────────────────────────────────
    stats = cargas.aggregate(
        total_cargas=Count('id'),
        total_litros=Sum('cantidad_litros'),
        promedio_litros=Avg('cantidad_litros'),
    )
    total_cargas = stats['total_cargas'] or 0
    total_litros = float(stats['total_litros'] or 0)
    promedio_litros = round(float(stats['promedio_litros'] or 0), 2)

    # Candados irregulares
    candados_irregulares = cargas.filter(
        estado_candado_anterior__in=['ALTERADO', 'VIOLADO', 'SIN_CANDADO']
    ).count()

    # ── Top 10 unidades ────────────────────────────────────────────────────────
    top_10 = list(
        CargaCombustible.objects
        .filter(fecha_hora_inicio__range=(fecha_inicio, fecha_fin), estado='COMPLETADO')
        .values('unidad__numero_economico', 'unidad__tipo')
        .annotate(total_litros=Sum('cantidad_litros'), num_cargas=Count('id'))
        .order_by('-total_litros')[:10]
    )

    # ── Unidades que NO cargaron ───────────────────────────────────────────────
    ids_que_cargaron = cargas.values_list('unidad_id', flat=True).distinct()
    no_cargaron = Unidad.objects.filter(activa=True).exclude(id__in=ids_que_cargaron)[:10]

    # ── Construcción del cuerpo ────────────────────────────────────────────────
    lineas = [
        f'REPORTE DE COMBUSTIBLE — {label}',
        f'Período: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}',
        '=' * 60,
        '',
        'RESUMEN GENERAL',
        '-' * 40,
        f'  Total cargas completadas : {total_cargas}',
        f'  Total litros cargados    : {total_litros:,.2f} L',
        f'  Promedio por carga       : {promedio_litros:,.2f} L',
        f'  Anomalías de candado     : {candados_irregulares}',
        '',
        'TOP 10 UNIDADES — MAYOR CONSUMO',
        '-' * 40,
    ]

    if top_10:
        for pos, u in enumerate(top_10, start=1):
            lineas.append(
                f'  {pos:2}. Eco. {u["unidad__numero_economico"]:<8} '
                f'({u["unidad__tipo"]:<8}) — '
                f'{float(u["total_litros"]):>8.2f} L  |  {u["num_cargas"]} carga(s)'
            )
    else:
        lineas.append('  Sin cargas en el período.')

    lineas += [
        '',
        f'UNIDADES ACTIVAS SIN CARGA EN EL PERÍODO (máx. 10)',
        '-' * 40,
    ]

    if no_cargaron.exists():
        for u in no_cargaron:
            lineas.append(f'  - Eco. {u.numero_economico}  —  {u.placa}  ({u.get_tipo_display()})')
    else:
        lineas.append('  Todas las unidades activas cargaron en el período.')

    lineas += [
        '',
        '=' * 60,
        'Se adjunta archivo Excel con el detalle completo de cargas.',
        '---',
        'Sistema BitacoraKasu — Transportes Kasu',
    ]

    return '\n'.join(lineas)


def enviar_reporte_combustible():
    """Función principal llamada por el scheduler."""
    cfg = settings.REPORTES_CONFIG.get('combustible', {})
    periodicidad = cfg.get('periodicidad', 'mensual')
    destinatarios = cfg.get('destinatarios', [])

    if not destinatarios:
        return

    fecha_inicio, fecha_fin, label = _get_periodo(periodicidad)

    try:
        cuerpo = generar_cuerpo_reporte(fecha_inicio, fecha_fin, label)
        excel_buffer = generar_excel_combustible(fecha_inicio, fecha_fin)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Error generando reporte combustible: {e}')
        return

    asunto = f'[Reporte] Combustible — {label}'
    nombre_archivo = f'reporte_combustible_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx'

    email = EmailMessage(
        subject=asunto,
        body=cuerpo,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=destinatarios,
    )
    email.attach(
        nombre_archivo,
        excel_buffer.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    try:
        email.send(fail_silently=False)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Error enviando reporte combustible: {e}')
