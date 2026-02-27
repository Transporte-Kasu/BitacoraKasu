"""
Reporte automático de salidas de almacén.
Genera cuerpo estadístico + Excel con todas las salidas del período.
"""
import io
from datetime import timedelta
from django.utils import timezone
from django.core.mail import EmailMessage
from django.conf import settings
from django.db.models import Sum, Count, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Período ──────────────────────────────────────────────────────────────────

def _get_periodo(periodicidad):
    """Calcula fecha_inicio y fecha_fin según la periodicidad."""
    ahora = timezone.now()
    if periodicidad == 'diario':
        fecha_inicio = (ahora - timedelta(days=1)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        fecha_fin = ahora
        label = f"Día {timezone.localtime(ahora - timedelta(days=1)).strftime('%d/%m/%Y')}"
    elif periodicidad == 'semanal':
        fecha_inicio = ahora - timedelta(days=7)
        fecha_fin = ahora
        fi_local = timezone.localtime(fecha_inicio)
        ff_local = timezone.localtime(fecha_fin)
        label = f"Semana {fi_local.strftime('%d/%m')} - {ff_local.strftime('%d/%m/%Y')}"
    else:  # mensual: mes anterior completo
        ahora_local = timezone.localtime(ahora)
        if ahora_local.month == 1:
            primer_dia_mes_ant = ahora_local.replace(
                year=ahora_local.year - 1, month=12, day=1,
                hour=0, minute=0, second=0, microsecond=0
            )
        else:
            primer_dia_mes_ant = ahora_local.replace(
                month=ahora_local.month - 1, day=1,
                hour=0, minute=0, second=0, microsecond=0
            )
        primer_dia_mes_actual = ahora_local.replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        )
        fecha_inicio = primer_dia_mes_ant.astimezone(timezone.utc)
        fecha_fin = (primer_dia_mes_actual - timedelta(seconds=1)).astimezone(timezone.utc)
        label = primer_dia_mes_ant.strftime('%B %Y').capitalize()
    return fecha_inicio, fecha_fin, label


# ── Estadísticas ──────────────────────────────────────────────────────────────

def _top10_productos(fecha_inicio, fecha_fin):
    """
    Top 10 productos más salidos en el período,
    combinando SalidaAlmacen + SalidaRapidaConsumible.
    """
    from modulos.almacen.models import ItemSalidaAlmacen, SalidaRapidaConsumible

    acum = {}

    # Salidas formales
    for r in (
        ItemSalidaAlmacen.objects
        .filter(salida__fecha_salida__range=(fecha_inicio, fecha_fin))
        .values('producto_almacen_id',
                'producto_almacen__descripcion',
                'producto_almacen__sku',
                'producto_almacen__unidad_medida')
        .annotate(total=Sum('cantidad_entregada'), veces=Count('id'))
    ):
        pid = r['producto_almacen_id']
        acum[pid] = {
            'descripcion': r['producto_almacen__descripcion'],
            'sku':         r['producto_almacen__sku'],
            'unidad':      r['producto_almacen__unidad_medida'],
            'total':       float(r['total']),
            'veces':       r['veces'],
        }

    # Salidas rápidas
    for r in (
        SalidaRapidaConsumible.objects
        .filter(fecha_salida__range=(fecha_inicio, fecha_fin))
        .values('producto_id',
                'producto__descripcion',
                'producto__sku',
                'producto__unidad_medida')
        .annotate(total=Sum('cantidad'), veces=Count('id'))
    ):
        pid = r['producto_id']
        if pid in acum:
            acum[pid]['total'] += float(r['total'])
            acum[pid]['veces'] += r['veces']
        else:
            acum[pid] = {
                'descripcion': r['producto__descripcion'],
                'sku':         r['producto__sku'],
                'unidad':      r['producto__unidad_medida'],
                'total':       float(r['total']),
                'veces':       r['veces'],
            }

    return sorted(acum.values(), key=lambda x: x['total'], reverse=True)[:10]


def _productos_sin_movimiento(dias=90):
    """
    Productos activos con stock > 0 que no han tenido ninguna salida
    en los últimos `dias` días.
    """
    from modulos.almacen.models import (
        ProductoAlmacen, ItemSalidaAlmacen, SalidaRapidaConsumible
    )
    fecha_limite = timezone.now() - timedelta(days=dias)

    ids_activos = set(
        ItemSalidaAlmacen.objects
        .filter(salida__fecha_salida__gte=fecha_limite)
        .values_list('producto_almacen_id', flat=True)
        .distinct()
    )
    ids_activos.update(
        SalidaRapidaConsumible.objects
        .filter(fecha_salida__gte=fecha_limite)
        .values_list('producto_id', flat=True)
        .distinct()
    )

    return list(
        ProductoAlmacen.objects
        .filter(activo=True, cantidad__gt=0)
        .exclude(id__in=ids_activos)
        .order_by('-cantidad')
        .values('descripcion', 'sku', 'cantidad', 'unidad_medida', 'categoria')
    )


def _productos_sin_salida_historico():
    """
    Productos activos con stock > 0 que han tenido entradas al almacén
    pero NUNCA han registrado una salida (en todo el historial).
    """
    from modulos.almacen.models import (
        ProductoAlmacen, ItemEntradaAlmacen,
        ItemSalidaAlmacen, SalidaRapidaConsumible
    )

    ids_con_entrada = set(
        ItemEntradaAlmacen.objects
        .values_list('producto_almacen_id', flat=True)
        .distinct()
    )
    ids_con_salida = set(
        ItemSalidaAlmacen.objects
        .values_list('producto_almacen_id', flat=True)
        .distinct()
    )
    ids_con_salida.update(
        SalidaRapidaConsumible.objects
        .values_list('producto_id', flat=True)
        .distinct()
    )

    solo_entradas = ids_con_entrada - ids_con_salida

    return list(
        ProductoAlmacen.objects
        .filter(activo=True, cantidad__gt=0, id__in=solo_entradas)
        .order_by('-cantidad')
        .values('descripcion', 'sku', 'cantidad', 'unidad_medida', 'categoria')
    )


# ── Cuerpo del correo ─────────────────────────────────────────────────────────

def generar_cuerpo_reporte(fecha_inicio, fecha_fin, label):
    """Genera el cuerpo de texto del correo con estadísticas."""
    from modulos.almacen.models import SalidaAlmacen, SalidaRapidaConsumible

    fi = timezone.localtime(fecha_inicio).strftime('%d/%m/%Y')
    ff = timezone.localtime(fecha_fin).strftime('%d/%m/%Y')

    # Totales del período
    total_salidas_formales = SalidaAlmacen.objects.filter(
        fecha_salida__range=(fecha_inicio, fecha_fin)
    ).count()
    total_rapidas = SalidaRapidaConsumible.objects.filter(
        fecha_salida__range=(fecha_inicio, fecha_fin)
    ).count()
    total_litros_rapidas = SalidaRapidaConsumible.objects.filter(
        fecha_salida__range=(fecha_inicio, fecha_fin)
    ).aggregate(t=Sum('cantidad'))['t'] or 0

    top10 = _top10_productos(fecha_inicio, fecha_fin)
    sin_movimiento = _productos_sin_movimiento(dias=90)
    sin_salida = _productos_sin_salida_historico()

    lineas = [
        f'REPORTE DE SALIDAS DE ALMACÉN — {label}',
        f'Período: {fi} al {ff}',
        '=' * 65,
        '',
        'RESUMEN DEL PERÍODO',
        '-' * 40,
        f'  Salidas formales (SOL/SAL)    : {total_salidas_formales}',
        f'  Salidas rápidas consumibles    : {total_rapidas}',
        f'  Total unidades salidas rápidas : {float(total_litros_rapidas):,.2f}',
        '',
        'TOP 10 PRODUCTOS MÁS SALIDOS EN EL PERÍODO',
        '-' * 40,
    ]

    if top10:
        for pos, p in enumerate(top10, start=1):
            lineas.append(
                f'  {pos:2}. {p["descripcion"][:40]:<40}  '
                f'{p["total"]:>8.2f} {p["unidad"]:<6}  '
                f'({p["veces"]} salida(s))  SKU: {p["sku"]}'
            )
    else:
        lineas.append('  Sin salidas registradas en el período.')

    lineas += [
        '',
        f'PRODUCTOS SIN MOVIMIENTO EN LOS ÚLTIMOS 90 DÍAS (con stock)',
        '-' * 40,
    ]
    if sin_movimiento:
        for p in sin_movimiento[:15]:
            lineas.append(
                f'  • {p["descripcion"][:45]:<45}  '
                f'Stock: {float(p["cantidad"]):>8.2f} {p["unidad_medida"]}  '
                f'[{p["categoria"]}]'
            )
        if len(sin_movimiento) > 15:
            lineas.append(f'  ... y {len(sin_movimiento) - 15} más (ver Excel)')
    else:
        lineas.append('  Todos los productos con stock han tenido movimiento reciente.')

    lineas += [
        '',
        'PRODUCTOS CON ENTRADAS PERO SIN NINGUNA SALIDA REGISTRADA',
        '-' * 40,
    ]
    if sin_salida:
        for p in sin_salida[:15]:
            lineas.append(
                f'  • {p["descripcion"][:45]:<45}  '
                f'Stock: {float(p["cantidad"]):>8.2f} {p["unidad_medida"]}  '
                f'[{p["categoria"]}]'
            )
        if len(sin_salida) > 15:
            lineas.append(f'  ... y {len(sin_salida) - 15} más (ver Excel)')
    else:
        lineas.append('  Todos los productos con entradas tienen al menos una salida.')

    lineas += [
        '',
        '=' * 65,
        'Se adjunta archivo Excel con el detalle completo.',
        '---',
        'Sistema BitacoraKasu — Transportes Kasu',
    ]

    return '\n'.join(lineas)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _fill_header():
    return PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

def _fill_subheader():
    return PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')

def _fill_zebra():
    return PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

def _font_header():
    return Font(color='FFFFFF', bold=True, size=10)

def _font_bold():
    return Font(bold=True)

def _aplicar_encabezado(ws, headers, row=1, fill=None):
    f = fill or _fill_header()
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = f
        c.font = _font_header()
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 28


def generar_excel_almacen(fecha_inicio, fecha_fin):
    """
    Genera Excel con 4 hojas:
      1. Salidas Almacen   — SalidaAlmacen + ItemSalidaAlmacen
      2. Salidas Rapidas   — SalidaRapidaConsumible
      3. Estadisticas      — Top 10, sin movimiento, sin salida
      4. Resumen           — Totales del período
    """
    from modulos.almacen.models import SalidaAlmacen, SalidaRapidaConsumible

    wb = openpyxl.Workbook()

    # ── Hoja 1: Salidas Almacén ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Salidas Almacen'
    headers1 = [
        'Folio Salida', 'Folio Solicitud', 'Tipo', 'Fecha Salida',
        'Entregado A', 'Entregado Por', 'Producto', 'SKU',
        'Cantidad', 'Unidad', 'Lote', 'Ubicación', 'Observaciones',
    ]
    _aplicar_encabezado(ws1, headers1)

    salidas = SalidaAlmacen.objects.select_related(
        'solicitud_salida', 'entregado_a', 'entregado_por'
    ).prefetch_related(
        'items__producto_almacen'
    ).filter(fecha_salida__range=(fecha_inicio, fecha_fin)).order_by('-fecha_salida')

    row = 2
    for salida in salidas:
        tipo = salida.solicitud_salida.get_tipo_display() if salida.solicitud_salida else ''
        folio_sol = salida.solicitud_salida.folio if salida.solicitud_salida else ''
        ent_a = salida.entregado_a.get_full_name() or salida.entregado_a.username
        ent_por = salida.entregado_por.get_full_name() or salida.entregado_por.username
        items = salida.items.all()
        filas = list(items) if items.exists() else [None]
        for item in filas:
            datos = [
                salida.folio, folio_sol, tipo,
                timezone.localtime(salida.fecha_salida).strftime('%d/%m/%Y %H:%M'),
                ent_a, ent_por,
                item.producto_almacen.descripcion if item else '',
                item.producto_almacen.sku if item else '',
                float(item.cantidad_entregada) if item else '',
                item.producto_almacen.unidad_medida if item else '',
                (item.lote or '') if item else '',
                (item.ubicacion_origen or '') if item else '',
                salida.observaciones or '',
            ]
            ws1.append(datos)
            if row % 2 == 0:
                for col in range(1, len(headers1) + 1):
                    ws1.cell(row=row, column=col).fill = _fill_zebra()
            row += 1

    for i, w in enumerate([16, 16, 22, 18, 22, 22, 38, 14, 10, 10, 14, 20, 30], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Salidas Rápidas ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Salidas Rapidas')
    headers2 = [
        'Folio', 'Fecha', 'Producto', 'SKU', 'Cantidad', 'Unidad',
        'Solicitante', 'Unidad Vehicular', 'Motivo', 'Entregado Por',
    ]
    _aplicar_encabezado(ws2, headers2)

    rapidas = SalidaRapidaConsumible.objects.select_related(
        'producto', 'unidad', 'entregado_por'
    ).filter(fecha_salida__range=(fecha_inicio, fecha_fin)).order_by('-fecha_salida')

    for row, sr in enumerate(rapidas, start=2):
        datos = [
            sr.folio,
            timezone.localtime(sr.fecha_salida).strftime('%d/%m/%Y %H:%M'),
            sr.producto.descripcion, sr.producto.sku,
            float(sr.cantidad), sr.producto.unidad_medida,
            sr.solicitante,
            sr.unidad.numero_economico if sr.unidad else '',
            sr.motivo or '',
            sr.entregado_por.get_full_name() or sr.entregado_por.username,
        ]
        ws2.append(datos)
        if row % 2 == 0:
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=col).fill = _fill_zebra()

    for i, w in enumerate([16, 18, 38, 14, 10, 10, 25, 16, 38, 25], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 3: Estadísticas ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('Estadisticas')
    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 22

    def titulo(ws, row, text, cols=5):
        c = ws.cell(row=row, column=1, value=text)
        c.fill = _fill_header()
        c.font = _font_header()
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        ws.row_dimensions[row].height = 22

    top10 = _top10_productos(fecha_inicio, fecha_fin)
    sin_mov = _productos_sin_movimiento(dias=90)
    sin_sal = _productos_sin_salida_historico()

    r = 1
    titulo(ws3, r, 'TOP 10 PRODUCTOS MÁS SALIDOS EN EL PERÍODO'); r += 1
    _aplicar_encabezado(
        ws3,
        ['Producto', 'SKU', 'Cantidad', 'Unidad', 'Salidas'],
        row=r, fill=_fill_subheader()
    ); r += 1

    if top10:
        for pos, p in enumerate(top10, start=1):
            ws3.cell(row=r, column=1, value=f'{pos}. {p["descripcion"]}')
            ws3.cell(row=r, column=2, value=p['sku'])
            ws3.cell(row=r, column=3, value=p['total'])
            ws3.cell(row=r, column=4, value=p['unidad'])
            ws3.cell(row=r, column=5, value=p['veces'])
            if r % 2 == 0:
                for col in range(1, 6):
                    ws3.cell(row=r, column=col).fill = _fill_zebra()
            r += 1
    else:
        ws3.cell(row=r, column=1, value='Sin salidas en el período.'); r += 1

    r += 1
    titulo(ws3, r, 'PRODUCTOS SIN MOVIMIENTO EN ÚLTIMOS 90 DÍAS (con stock)'); r += 1
    _aplicar_encabezado(
        ws3,
        ['Producto', 'SKU', 'Stock', 'Unidad', 'Categoría'],
        row=r, fill=_fill_subheader()
    ); r += 1

    if sin_mov:
        for p in sin_mov:
            ws3.cell(row=r, column=1, value=p['descripcion'])
            ws3.cell(row=r, column=2, value=p['sku'])
            ws3.cell(row=r, column=3, value=float(p['cantidad']))
            ws3.cell(row=r, column=4, value=p['unidad_medida'])
            ws3.cell(row=r, column=5, value=p['categoria'])
            if r % 2 == 0:
                for col in range(1, 6):
                    ws3.cell(row=r, column=col).fill = _fill_zebra()
            r += 1
    else:
        ws3.cell(row=r, column=1, value='Todos los productos con stock han tenido movimiento.'); r += 1

    r += 1
    titulo(ws3, r, 'PRODUCTOS CON ENTRADAS PERO SIN SALIDAS (historial completo)'); r += 1
    _aplicar_encabezado(
        ws3,
        ['Producto', 'SKU', 'Stock', 'Unidad', 'Categoría'],
        row=r, fill=_fill_subheader()
    ); r += 1

    if sin_sal:
        for p in sin_sal:
            ws3.cell(row=r, column=1, value=p['descripcion'])
            ws3.cell(row=r, column=2, value=p['sku'])
            ws3.cell(row=r, column=3, value=float(p['cantidad']))
            ws3.cell(row=r, column=4, value=p['unidad_medida'])
            ws3.cell(row=r, column=5, value=p['categoria'])
            if r % 2 == 0:
                for col in range(1, 6):
                    ws3.cell(row=r, column=col).fill = _fill_zebra()
            r += 1
    else:
        ws3.cell(row=r, column=1, value='Todos los productos tienen al menos una salida registrada.'); r += 1

    # ── Hoja 4: Resumen ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Resumen')
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 20

    def dato(ws, row, label, valor):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = _font_bold()
        ws.cell(row=row, column=2, value=valor)

    fi_str = timezone.localtime(fecha_inicio).strftime('%d/%m/%Y')
    ff_str = timezone.localtime(fecha_fin).strftime('%d/%m/%Y')

    total_salidas = salidas.count()
    total_items_entregados = sum(s.items.count() for s in salidas)
    total_rapidas_cnt = rapidas.count()
    total_cant_rapidas = rapidas.aggregate(t=Sum('cantidad'))['t'] or 0

    r2 = 1
    titulo(ws4, r2, 'RESUMEN DE SALIDAS DE ALMACÉN'); r2 += 1
    dato(ws4, r2, 'Período', f'{fi_str} — {ff_str}'); r2 += 1
    r2 += 1
    titulo(ws4, r2, 'Salidas Formales (SOL/SAL)'); r2 += 1
    dato(ws4, r2, 'Total salidas registradas', total_salidas); r2 += 1
    dato(ws4, r2, 'Total ítems entregados', total_items_entregados); r2 += 1
    r2 += 1
    titulo(ws4, r2, 'Salidas Rápidas de Consumibles'); r2 += 1
    dato(ws4, r2, 'Total salidas rápidas', total_rapidas_cnt); r2 += 1
    dato(ws4, r2, 'Cantidad total entregada', float(total_cant_rapidas)); r2 += 1
    r2 += 1
    titulo(ws4, r2, 'Estadísticas de Inventario'); r2 += 1
    dato(ws4, r2, 'Productos sin movimiento (90 días)', len(sin_mov)); r2 += 1
    dato(ws4, r2, 'Productos sin salida (historial)', len(sin_sal)); r2 += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Envío ─────────────────────────────────────────────────────────────────────

def enviar_reporte_almacen():
    """Función principal llamada por el scheduler."""
    cfg = settings.REPORTES_CONFIG.get('almacen', {})
    periodicidad = cfg.get('periodicidad', 'semanal')
    destinatarios = cfg.get('destinatarios', [])

    if not destinatarios:
        return

    fecha_inicio, fecha_fin, label = _get_periodo(periodicidad)

    try:
        cuerpo = generar_cuerpo_reporte(fecha_inicio, fecha_fin, label)
        excel_buffer = generar_excel_almacen(fecha_inicio, fecha_fin)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Error generando reporte almacén: {e}')
        return

    asunto = f'[Reporte] Salidas de Almacén — {label}'
    nombre_archivo = (
        f'reporte_almacen_'
        f'{timezone.localtime(fecha_inicio).strftime("%Y%m%d")}_'
        f'{timezone.localtime(fecha_fin).strftime("%Y%m%d")}.xlsx'
    )

    email = EmailMessage(
        subject=asunto,
        body=cuerpo,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=destinatarios,
    )
    email.attach(
        nombre_archivo,
        excel_buffer.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    try:
        email.send(fail_silently=False)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Error enviando reporte almacén: {e}')
