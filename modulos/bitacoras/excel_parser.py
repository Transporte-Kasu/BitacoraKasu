import re
import openpyxl
from datetime import datetime, timedelta


def _extract_cp(text):
    if not text:
        return ''
    match = re.search(r'[Cc]\.?[Pp]\.?\s*(\d{5})', str(text))
    return match.group(1) if match else ''


def _parse_fecha_entrega(raw):
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if not m:
        return None
    day, month, year = m.groups()
    return datetime(int(year), int(month), int(day))


def parse_confirmacion_excel(file_obj, hora_salida_str, hora_carga_str, tipo_contenedor):
    """
    Parses CONFIRMACION_SERVICIOS.xlsx.
    Returns list of dicts, one per viaje (FULL: pairs merged; SENCILLO: single row).
    Columns: A=fecha entrega, B=contenedor, C=custodia, D=dir carta porte,
             E=dir entrega, F=modalidad, G=contacto bodega, H=codigo SAT,
             I=mercancia, J=unidad medida, K=cantidad, L=pesos(kg), M=pedimento
    """
    hora_sal_h, hora_sal_m = (int(x) for x in hora_salida_str.split(':'))
    hora_car_h, hora_car_m = (int(x) for x in hora_carga_str.split(':'))

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[1] and str(row[1]).strip().upper() == 'CONTENEDOR':
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise ValueError("No se encontró la fila de encabezados (columna B debe decir CONTENEDOR).")

    viajes = []
    current = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue

        contenedor = row[1]
        if not contenedor:
            continue

        contenedor    = str(contenedor).strip()
        fecha_raw        = row[0]
        custodia         = row[2]
        dir_carta_porte  = row[3]
        dir_entrega      = row[4]
        modalidad_raw    = row[5]
        contacto         = row[6]
        mercancia        = row[8]
        pesos_kg         = row[11]
        pedimento        = row[12]

        peso_tons = round(float(pesos_kg) / 1000, 3) if pesos_kg else None

        if modalidad_raw is not None:
            raw_str = str(modalidad_raw).strip()
            raw_up  = raw_str.upper()

            if 'FULL' in raw_up:
                modalidad     = 'FULL'
                instrucciones = '' if raw_up == 'FULL 1X1' else raw_str
            elif 'SENCILLO' in raw_up:
                modalidad     = 'SENCILLO'
                instrucciones = ''
            else:
                modalidad     = raw_up
                instrucciones = ''

            obs_parts = []
            if instrucciones:
                obs_parts.append(instrucciones)
            if custodia:
                obs_parts.append(f'Custodia: {custodia}')
            if contacto:
                obs_parts.append(f'Contacto: {contacto}')
            if mercancia:
                obs_parts.append(f'Mercancía: {mercancia}')
            if pedimento:
                obs_parts.append(f'Pedimento: {pedimento}')

            fecha_entrega          = _parse_fecha_entrega(fecha_raw)
            cp                     = _extract_cp(dir_entrega)
            destino                = str(dir_entrega).strip() if dir_entrega else ''
            domicilio_carta_porte  = str(dir_carta_porte).strip() if dir_carta_porte else ''

            fecha_salida = fecha_carga = None
            if fecha_entrega:
                dia_ant      = fecha_entrega - timedelta(days=1)
                fecha_salida = dia_ant.replace(hour=hora_sal_h, minute=hora_sal_m, second=0, microsecond=0)
                fecha_carga  = dia_ant.replace(hour=hora_car_h, minute=hora_car_m, second=0, microsecond=0)

            current = {
                'modalidad':             modalidad,
                'contenedor':            contenedor,
                'contenedor_2':          '',
                'peso':                  str(peso_tons) if peso_tons is not None else '',
                'peso_2':                '',
                'destino':               destino,
                'cp_destino':            cp,
                'domicilio_carta_porte': domicilio_carta_porte,
                'cp_faltante':           not cp,
                'fecha_entrega_display': fecha_entrega.strftime('%d/%m/%Y') if fecha_entrega else '',
                'fecha_salida':          fecha_salida.strftime('%Y-%m-%dT%H:%M') if fecha_salida else '',
                'fecha_carga':           fecha_carga.strftime('%Y-%m-%dT%H:%M') if fecha_carga else '',
                'observaciones':         '\n'.join(obs_parts),
                'tipo_contenedor':       tipo_contenedor,
            }
            viajes.append(current)

        else:
            # Segunda línea de un FULL
            if current and current['modalidad'] == 'FULL':
                current['contenedor_2'] = contenedor
                current['peso_2']       = str(peso_tons) if peso_tons is not None else ''
                if pedimento:
                    sep = '\n' if current['observaciones'] else ''
                    current['observaciones'] += f'{sep}Pedimento 2: {pedimento}'

    return viajes
