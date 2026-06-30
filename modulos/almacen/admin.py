from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from django.utils import timezone
from .models import (
    ProductoAlmacen, EntradaAlmacen, ItemEntradaAlmacen,
    SolicitudSalida, ItemSolicitudSalida, SalidaAlmacen,
    ItemSalidaAlmacen, MovimientoAlmacen, AlertaStock,
    SalidaRapidaConsumible, AuditoriaAlmacen
)


class ItemEntradaAlmacenInline(admin.TabularInline):
    model = ItemEntradaAlmacen
    extra = 1
    fields = ['producto_almacen', 'cantidad', 'costo_unitario', 'lote', 'ubicacion_asignada']


class ItemSolicitudSalidaInline(admin.TabularInline):
    model = ItemSolicitudSalida
    extra = 1
    fields = ['producto_almacen', 'cantidad_solicitada', 'cantidad_entregada']
    readonly_fields = ['cantidad_entregada']


class ItemSalidaAlmacenInline(admin.TabularInline):
    model = ItemSalidaAlmacen
    extra = 0
    fields = ['producto_almacen', 'cantidad_entregada', 'lote', 'ubicacion_origen']


@admin.register(ProductoAlmacen)
class ProductoAlmacenAdmin(admin.ModelAdmin):
    list_display = [
        'sku', 'descripcion', 'categoria', 'subcategoria', 'cantidad',
        'unidad_medida', 'stock_minimo', 'costo_unitario', 'es_consumible', 'activo'
    ]
    list_filter = ['categoria', 'activo', 'tiene_caducidad', 'es_consumible']
    search_fields = ['sku', 'descripcion', 'codigo_barras']
    readonly_fields = ['fecha_registro', 'fecha_actualizacion']
    fieldsets = (
        ('Información Básica', {
            'fields': ('categoria', 'subcategoria', 'sku', 'codigo_barras', 'descripcion', 'imagen')
        }),
        ('Ubicación y Stock', {
            'fields': ('localidad', 'cantidad', 'unidad_medida', 'stock_minimo', 'stock_maximo')
        }),
        ('Costos', {
            'fields': ('costo_unitario',)
        }),
        ('Caducidad', {
            'fields': ('tiene_caducidad', 'fecha_caducidad')
        }),
        ('Relaciones', {
            'fields': ('producto_compra', 'proveedor_principal', 'tiempo_reorden_dias')
        }),
        ('Tipo de Producto', {
            'fields': ('es_consumible',)
        }),
        ('Metadata', {
            'fields': ('notas', 'activo', 'fecha_registro', 'fecha_actualizacion')
        }),
    )


@admin.register(EntradaAlmacen)
class EntradaAlmacenAdmin(admin.ModelAdmin):
    list_display = [
        'folio', 'tipo', 'fecha_entrada', 'recibido_por', 'total_items'
    ]
    list_filter = ['tipo', 'fecha_entrada']
    search_fields = ['folio', 'factura_numero']
    readonly_fields = ['folio']
    inlines = [ItemEntradaAlmacenInline]
    fieldsets = (
        ('Información Básica', {
            'fields': ('tipo', 'folio', 'fecha_entrada', 'recibido_por')
        }),
        ('Referencias', {
            'fields': ('orden_compra', 'orden_trabajo', 'recepcion_almacen_compras')
        }),
        ('Factura', {
            'fields': ('factura_numero', 'factura_archivo')
        }),
        ('Costos Adicionales', {
            'fields': ('costo_envio', 'costo_adicional')
        }),
        ('Observaciones', {
            'fields': ('observaciones',)
        }),
    )


@admin.register(ItemEntradaAlmacen)
class ItemEntradaAlmacenAdmin(admin.ModelAdmin):
    list_display = [
        'entrada', 'producto_almacen', 'cantidad', 'costo_unitario', 'costo_total'
    ]
    list_filter = ['entrada__tipo']
    search_fields = ['entrada__folio', 'producto_almacen__sku', 'producto_almacen__descripcion']


@admin.register(SolicitudSalida)
class SolicitudSalidaAdmin(admin.ModelAdmin):
    list_display = [
        'folio', 'tipo', 'estado', 'solicitante', 'fecha_solicitud',
        'requiere_autorizacion', 'autorizado_por'
    ]
    list_filter = ['tipo', 'estado', 'requiere_autorizacion', 'fecha_solicitud']
    search_fields = ['folio', 'solicitante__username']
    readonly_fields = ['folio', 'fecha_solicitud', 'fecha_autorizacion']
    inlines = [ItemSolicitudSalidaInline]
    fieldsets = (
        ('Información Básica', {
            'fields': ('tipo', 'folio', 'solicitante', 'fecha_solicitud', 'estado')
        }),
        ('Referencias', {
            'fields': ('orden_trabajo',)
        }),
        ('Justificación', {
            'fields': ('justificacion',)
        }),
        ('Autorización', {
            'fields': (
                'requiere_autorizacion', 'autorizado_por', 'fecha_autorizacion',
                'comentarios_autorizacion'
            )
        }),
    )


@admin.register(ItemSolicitudSalida)
class ItemSolicitudSalidaAdmin(admin.ModelAdmin):
    list_display = [
        'solicitud', 'producto_almacen', 'cantidad_solicitada',
        'cantidad_entregada', 'cantidad_pendiente'
    ]
    list_filter = ['solicitud__estado']
    search_fields = ['solicitud__folio', 'producto_almacen__sku']


@admin.register(SalidaAlmacen)
class SalidaAlmacenAdmin(admin.ModelAdmin):
    list_display = [
        'folio', 'solicitud_salida', 'fecha_salida', 'entregado_a', 'entregado_por'
    ]
    list_filter = ['fecha_salida']
    search_fields = ['folio', 'solicitud_salida__folio']
    readonly_fields = ['folio']
    inlines = [ItemSalidaAlmacenInline]


@admin.register(ItemSalidaAlmacen)
class ItemSalidaAlmacenAdmin(admin.ModelAdmin):
    list_display = [
        'salida', 'producto_almacen', 'cantidad_entregada', 'lote'
    ]
    search_fields = ['salida__folio', 'producto_almacen__sku']


@admin.register(MovimientoAlmacen)
class MovimientoAlmacenAdmin(admin.ModelAdmin):
    list_display = [
        'fecha_movimiento', 'tipo', 'producto_almacen', 'cantidad',
        'cantidad_anterior', 'cantidad_posterior', 'usuario'
    ]
    list_filter = ['tipo', 'fecha_movimiento']
    search_fields = ['producto_almacen__sku', 'producto_almacen__descripcion']
    readonly_fields = ['fecha_movimiento']
    date_hierarchy = 'fecha_movimiento'


@admin.register(AlertaStock)
class AlertaStockAdmin(admin.ModelAdmin):
    list_display = [
        'tipo_alerta', 'producto_almacen', 'fecha_generacion', 'resuelta',
        'resuelta_por', 'fecha_resolucion'
    ]
    list_filter = ['tipo_alerta', 'resuelta', 'fecha_generacion']
    search_fields = ['producto_almacen__sku', 'producto_almacen__descripcion', 'mensaje']
    readonly_fields = ['fecha_generacion', 'fecha_resolucion']
    date_hierarchy = 'fecha_generacion'


@admin.register(SalidaRapidaConsumible)
class SalidaRapidaConsumibleAdmin(admin.ModelAdmin):
    list_display = [
        'folio', 'producto', 'cantidad', 'solicitante',
        'entregado_por', 'fecha_salida'
    ]
    list_filter = ['fecha_salida', 'producto']
    search_fields = ['folio', 'producto__sku', 'producto__descripcion', 'solicitante']
    readonly_fields = ['folio']
    date_hierarchy = 'fecha_salida'


ACCION_COLORES = {
    'CREAR':     '#16a34a',
    'EDITAR':    '#2563eb',
    'ELIMINAR':  '#dc2626',
    'AUTORIZAR': '#0891b2',
    'RECHAZAR':  '#ea580c',
    'ENTREGAR':  '#7c3aed',
    'CANCELAR':  '#6b7280',
}


@admin.register(AuditoriaAlmacen)
class AuditoriaAlmacenAdmin(admin.ModelAdmin):
    change_list_template = 'admin/almacen/auditoriaalmacen/change_list.html'
    change_form_template = 'admin/almacen/auditoriaalmacen/change_form.html'

    list_display = ['fecha', 'usuario', 'badge_accion', 'modelo', 'objeto_str', 'ip_address']
    list_filter = ['accion', 'modelo', 'usuario', 'fecha']
    search_fields = ['usuario__username', 'usuario__first_name', 'objeto_str', 'modelo']
    readonly_fields = [
        'fecha', 'usuario', 'accion', 'modelo', 'objeto_id',
        'objeto_str', 'valores_anteriores', 'valores_nuevos', 'ip_address',
    ]
    date_hierarchy = 'fecha'
    ordering = ['-fecha']

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_module_perms(self, request):
        return request.user.is_superuser

    @admin.display(description='Acción')
    def badge_accion(self, obj):
        from django.utils.html import format_html
        color = ACCION_COLORES.get(obj.accion, '#6b7280')
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;">{}</span>',
            color, obj.get_accion_display()
        )

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('exportar-excel/', self.admin_site.admin_view(self.exportar_excel),
                 name='almacen_auditoriaalmacen_exportar_excel'),
        ]
        return custom + urls

    def exportar_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        qs = AuditoriaAlmacen.objects.select_related('usuario').order_by('-fecha')

        # Aplicar filtros de la querystring
        accion = request.GET.get('accion')
        modelo = request.GET.get('modelo')
        usuario = request.GET.get('usuario')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        if accion:
            qs = qs.filter(accion=accion)
        if modelo:
            qs = qs.filter(modelo=modelo)
        if usuario:
            qs = qs.filter(usuario__username=usuario)
        if fecha_desde:
            qs = qs.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__date__lte=fecha_hasta)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Auditoría Almacén'

        header_fill = PatternFill('solid', fgColor='1E3A5F')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ['Fecha', 'Usuario', 'Acción', 'Modelo', 'Objeto', 'Campo', 'Valor Anterior', 'Valor Nuevo', 'IP']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.row_dimensions[1].height = 20
        col_widths = [18, 15, 12, 20, 30, 22, 22, 22, 15]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.freeze_panes = 'A2'

        row = 2
        for registro in qs:
            antes = registro.valores_anteriores or {}
            despues = registro.valores_nuevos or {}
            usuario_str = registro.usuario.username if registro.usuario else '—'
            fecha_str = timezone.localtime(registro.fecha).strftime('%d/%m/%Y %H:%M')

            if registro.accion == 'CREAR':
                campos = [(k, '—', str(v) if v is not None else '—') for k, v in despues.items()]
            elif registro.accion == 'ELIMINAR':
                campos = [(k, str(v) if v is not None else '—', '—') for k, v in antes.items()]
            else:
                campos = [
                    (k, str(antes.get(k, '')) if antes.get(k) is not None else '—',
                     str(despues.get(k, '')) if despues.get(k) is not None else '—')
                    for k in despues
                    if str(antes.get(k, '')) != str(despues.get(k, ''))
                ]
                if not campos:
                    campos = [('(sin cambios)', '—', '—')]

            for campo, ant, nvo in campos:
                ws.cell(row=row, column=1, value=fecha_str).border = border
                ws.cell(row=row, column=2, value=usuario_str).border = border
                ws.cell(row=row, column=3, value=registro.get_accion_display()).border = border
                ws.cell(row=row, column=4, value=registro.modelo).border = border
                ws.cell(row=row, column=5, value=registro.objeto_str).border = border
                ws.cell(row=row, column=6, value=campo).border = border
                ws.cell(row=row, column=7, value=ant).border = border
                ws.cell(row=row, column=8, value=nvo).border = border
                ws.cell(row=row, column=9, value=registro.ip_address or '—').border = border
                row += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_hoy = timezone.now().strftime('%Y%m%d')
        response['Content-Disposition'] = f'attachment; filename="auditoria_almacen_{fecha_hoy}.xlsx"'
        wb.save(response)
        return response
