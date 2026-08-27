from datetime import timedelta, date
from decimal import Decimal

from django.contrib.auth.models import User
from django.test import TestCase
from django.utils import timezone

from modulos.reportes.models import ConfiguracionReporte
from modulos.reportes.generadores.flota import generar_vigencias_flota
from modulos.reportes.generadores.almacen import generar_analisis_integral
from modulos.reportes.generadores.unidades import generar_balanza_utilidad, GENERADORES
from modulos.equipos.models import Equipo
from modulos.dollys.models import Dolly
from modulos.caja_seca.models import CajaSeca
from modulos.almacen.models import (
    ProductoAlmacen, AsignacionDirectaAlmacen, AsignacionSalida,
    ItemAsignacionSalida, EntradaAlmacen, AuditoriaAlmacen,
)
from modulos.unidades.models import Unidad
from modulos.bitacoras.models import BitacoraViaje
from modulos.combustible.models import Despachador, CargaCombustible
from modulos.operadores.models import Operador


class EsDebidoMensualTests(TestCase):
    def _config(self, dia_mes, ultimo_envio):
        return ConfiguracionReporte(
            nombre='Test', modulo='FLOTA', tipo_reporte='FLOTA_VIGENCIAS',
            frecuencia='MENSUAL', dia_mes=dia_mes,
            destinatarios='a@b.com', ultimo_envio=ultimo_envio,
        )

    def test_no_dispara_antes_del_dia_configurado(self):
        hoy = timezone.now()
        mes_pasado = (hoy.replace(day=1) - timedelta(days=1)).replace(day=1)
        config = self._config(dia_mes=15, ultimo_envio=mes_pasado)
        if hoy.day < 15:
            self.assertFalse(config.es_debido())

    def test_dispara_en_o_despues_del_dia_configurado(self):
        hoy = timezone.now()
        if hoy.day < 15:
            self.skipTest('Solo verificable a partir del día 15 del mes actual')
        mes_pasado = (hoy.replace(day=1) - timedelta(days=1)).replace(day=1)
        config = self._config(dia_mes=15, ultimo_envio=mes_pasado)
        self.assertTrue(config.es_debido())

    def test_no_dispara_en_el_mismo_mes_del_ultimo_envio(self):
        hoy = timezone.now()
        config = self._config(dia_mes=1, ultimo_envio=hoy)
        self.assertFalse(config.es_debido())

    def test_sin_dia_mes_se_comporta_como_dia_1(self):
        hoy = timezone.now()
        mes_pasado = (hoy.replace(day=1) - timedelta(days=1)).replace(day=1)
        config = self._config(dia_mes=None, ultimo_envio=mes_pasado)
        self.assertTrue(config.es_debido())


class GenerarVigenciasFlotaTests(TestCase):
    def setUp(self):
        hoy = timezone.now().date()
        Equipo.objects.create(
            numero_economico='CHASIS-VENCIDO', numero_serie='S1',
            vigencia_doble_articulado=hoy - timedelta(days=5),
        )
        Equipo.objects.create(
            numero_economico='CHASIS-POR-VENCER', numero_serie='S2',
            vigencia_doble_articulado=hoy + timedelta(days=10),
        )
        Equipo.objects.create(
            numero_economico='CHASIS-VIGENTE', numero_serie='S3',
            vigencia_doble_articulado=hoy + timedelta(days=200),
        )
        Equipo.objects.create(
            numero_economico='CHASIS-SIN-DATO', numero_serie='S4',
        )
        Equipo.objects.create(
            numero_economico='CHASIS-INACTIVO', numero_serie='S5', activo=False,
            vigencia_doble_articulado=hoy - timedelta(days=100),
        )
        Dolly.objects.create(numero_economico='DOLLY-1', numero_serie='D1')
        CajaSeca.objects.create(numero_economico='CAJA-1', numero_serie='C1')

    def test_resumen_clasifica_equipos_correctamente(self):
        datos = generar_vigencias_flota(timezone.now().date(), timezone.now().date())
        resumen = datos['resumen']
        self.assertEqual(resumen['total_equipos'], 4)  # excluye el inactivo
        self.assertEqual(resumen['equipos_vencidos'], 1)
        self.assertEqual(resumen['equipos_por_vencer_30d'], 1)
        self.assertEqual(resumen['equipos_sin_dato'], 1)
        self.assertEqual(resumen['total_dollys'], 1)
        self.assertEqual(resumen['total_cajas_secas'], 1)

    def test_filas_incluye_los_tres_tipos_de_equipo(self):
        datos = generar_vigencias_flota(timezone.now().date(), timezone.now().date())
        tipos = {f['tipo_equipo'] for f in datos['filas']}
        self.assertEqual(tipos, {'Equipo', 'Dolly', 'Caja Seca'})

    def test_vencidos_aparecen_antes_que_vigentes(self):
        datos = generar_vigencias_flota(timezone.now().date(), timezone.now().date())
        estados = [f['estado_vigencia'] for f in datos['filas'] if f['tipo_equipo'] == 'Equipo']
        self.assertEqual(estados[0], 'VENCIDO')


class GenerarAnalisisIntegralAlmacenTests(TestCase):
    def setUp(self):
        self.hoy = timezone.now().date()

        self.mecanico1 = User.objects.create_user(
            username='mecanico1', first_name='Juan', last_name='Perez', password='12345'
        )
        self.mecanico2 = User.objects.create_user(
            username='mecanico2', first_name='Ana', last_name='Lopez', password='12345'
        )

        self.unidad = Unidad.objects.create(
            numero_economico='U-100', placa='ABC123', tipo='LOCAL', año=2020,
            capacidad_combustible=Decimal('200.00'), rendimiento_esperado=Decimal('3.50'),
        )

        self.producto = ProductoAlmacen.objects.create(
            categoria='Refacciones', sku='FOC-001', descripcion='Foco delantero',
            localidad='Pasillo A1', cantidad=Decimal('50.00'), unidad_medida='Pieza',
            stock_minimo=Decimal('5.00'), costo_unitario=Decimal('80.00'), activo=True,
        )

        AsignacionDirectaAlmacen.objects.create(
            producto=self.producto, unidad=self.unidad, cantidad=Decimal('2.00'),
            motivo='Cambio de foco', entregado_por=self.mecanico1,
            fecha_asignacion=timezone.now(),
        )

        asignacion_salida = AsignacionSalida.objects.create(
            solicitante='Juan Perez', tipo_destino='UNIDAD', unidad=self.unidad,
            justificacion='Reparación urgente', entregado_por=self.mecanico1,
        )
        ItemAsignacionSalida.objects.create(
            asignacion=asignacion_salida, producto=self.producto, cantidad=Decimal('3.00'),
        )

        EntradaAlmacen.objects.create(tipo='FACTURA', recibido_por=self.mecanico2)
        EntradaAlmacen.objects.create(tipo='ENTRADA_DIRECTA', recibido_por=self.mecanico2)

        # Los signals de auditoría generan eventos automáticos (usuario=None) para
        # los objetos creados arriba; se limpian para que el escenario de auditoría
        # de este test sea determinista.
        AuditoriaAlmacen.objects.all().delete()

        for _ in range(6):
            AuditoriaAlmacen.objects.create(
                usuario=self.mecanico1, accion='ELIMINAR', modelo='ProductoAlmacen',
                objeto_id='1', objeto_str='Producto de prueba',
            )
        AuditoriaAlmacen.objects.create(
            usuario=self.mecanico2, accion='CREAR', modelo='EntradaAlmacen',
            objeto_id='1', objeto_str='Entrada de prueba',
        )

    def test_resumen_cuenta_asignaciones_y_entradas(self):
        datos = generar_analisis_integral(self.hoy, self.hoy)
        resumen = datos['resumen']
        self.assertEqual(resumen['total_asignaciones_directas'], 1)
        self.assertEqual(resumen['total_asignaciones_salida'], 1)
        self.assertEqual(resumen['total_items_asignados'], 5.0)
        self.assertEqual(resumen['total_entradas'], 2)
        self.assertEqual(resumen['entradas_por_tipo']['Producto Nuevo desde Factura'], 1)
        self.assertEqual(resumen['entradas_por_tipo']['Entrada Directa'], 1)

    def test_top_destinos_incluye_la_unidad(self):
        datos = generar_analisis_integral(self.hoy, self.hoy)
        destinos = {d['destino'] for d in datos['top_destinos']}
        self.assertIn(str(self.unidad), destinos)

    def test_auditoria_detecta_concentracion_y_eliminaciones(self):
        datos = generar_analisis_integral(self.hoy, self.hoy)
        resumen = datos['resumen']
        self.assertEqual(resumen['total_eventos_auditoria'], 7)
        self.assertEqual(resumen['auditoria_por_accion']['Eliminar'], 6)
        self.assertEqual(len(resumen['alertas_auditoria']), 2)
        self.assertIn('Juan Perez', resumen['alertas_auditoria'][0])
        self.assertIn('eliminaciones', resumen['alertas_auditoria'][1].lower())

    def test_tablas_para_excel_incluye_las_tres_secciones(self):
        datos = generar_analisis_integral(self.hoy, self.hoy)
        self.assertEqual(set(datos['tablas'].keys()), {'Asignaciones', 'Entradas', 'Auditoria'})
        self.assertEqual(len(datos['tablas']['Asignaciones']), 2)


class PromptAnalisisIntegralAlmacenTests(TestCase):
    def test_prompt_incluye_kpis_principales(self):
        from modulos.reportes.generadores.narrativa import _prompt_almacen_analisis_integral

        resumen = {
            'total_asignaciones_directas': 3, 'total_asignaciones_salida': 2,
            'total_items_asignados': 12.0, 'total_entradas': 5,
            'entradas_por_tipo': {'Producto Nuevo desde Factura': 5},
            'valor_total_entradas': 1500.0, 'total_eventos_auditoria': 4,
            'auditoria_por_accion': {'Crear': 4},
            'alertas_auditoria': [
                'El usuario Juan Perez concentra el 90% de la actividad de auditoría del período.'
            ],
        }
        datos = {
            'top_destinos': [{'destino': 'Unidad U-100', 'cantidad_total': 8.0}],
            'top_usuarios_auditoria': [{'usuario': 'Juan Perez', 'total_eventos': 4}],
        }
        prompt, max_tokens = _prompt_almacen_analisis_integral(
            resumen, datos, '2026-07-01', '2026-07-22'
        )
        self.assertIn('Unidad U-100', prompt)
        self.assertIn('Juan Perez', prompt)
        self.assertIn('Producto Nuevo desde Factura', prompt)
        self.assertEqual(max_tokens, 600)


class GenerarExcelTablasTests(TestCase):
    def test_genera_una_hoja_por_tabla(self):
        import openpyxl
        from io import BytesIO
        from modulos.reportes.management.commands.generar_reportes import _generar_excel

        datos = {
            'titulo': 'Análisis Integral de Almacén',
            'resumen': {
                'total_entradas': 2,
                'entradas_por_tipo': {'Producto Nuevo desde Factura': 2},
                'alertas_auditoria': ['El usuario X concentra actividad.'],
            },
            'tablas': {
                'Asignaciones': [{'folio': 'ADI-1', 'cantidad': 2.0}],
                'Entradas': [{'folio': 'ENT-1', 'costo_total_entrada': 100.0}],
                'Auditoria': [{'usuario': 'Juan', 'total_eventos': 3}],
            },
        }
        contenido = _generar_excel(datos)
        wb = openpyxl.load_workbook(BytesIO(contenido))
        self.assertEqual(set(wb.sheetnames), {'Asignaciones', 'Entradas', 'Auditoria', 'Resumen'})
        ws = wb['Asignaciones']
        self.assertEqual(ws['A1'].value, 'Folio')
        self.assertEqual(ws['A2'].value, 'ADI-1')

    def test_resumen_con_dict_y_listas_no_lanza_error(self):
        import openpyxl
        from io import BytesIO
        from modulos.reportes.management.commands.generar_reportes import _generar_excel

        datos = {
            'titulo': 'Reporte', 'filas': [{'sku': 'A1', 'cantidad': 1}],
            'resumen': {'entradas_por_tipo': {'Factura': 2}, 'alertas_auditoria': ['x']},
        }
        contenido = _generar_excel(datos)
        wb = openpyxl.load_workbook(BytesIO(contenido))
        self.assertIn('Resumen', wb.sheetnames)


class EmailTemplateAnalisisIntegralTests(TestCase):
    def test_render_incluye_secciones_principales(self):
        from django.template.loader import render_to_string

        datos = {
            'tipo': 'ALMACEN_ANALISIS_INTEGRAL',
            'titulo': 'Análisis Integral de Almacén',
            'periodo_inicio': '2026-07-01', 'periodo_fin': '2026-07-22',
            'generado_en': '2026-07-22T09:00:00',
            'resumen': {
                'total_asignaciones_directas': 3, 'total_asignaciones_salida': 2,
                'total_entradas': 5, 'valor_total_entradas': 1500.0,
                'total_eventos_auditoria': 7,
                'entradas_por_tipo': {'Producto Nuevo desde Factura': 5},
                'alertas_auditoria': [
                    'El usuario Juan Perez concentra el 86% de la actividad de auditoría del período.'
                ],
            },
            'top_destinos': [{'destino': 'Unidad U-100', 'cantidad_total': 8.0}],
            'top_usuarios_auditoria': [{'usuario': 'Juan Perez', 'total_eventos': 6}],
        }
        html = render_to_string(
            'reportes/email/reporte_base.html', {'datos': datos, 'config': None, 'narrativa': ''}
        )
        self.assertIn('Unidad U-100', html)
        self.assertIn('Producto Nuevo desde Factura', html)
        self.assertIn('Juan Perez', html)
        self.assertIn('concentra el 86%', html)


class GenerarBalanzaUtilidadTests(TestCase):
    def setUp(self):
        self.unidad_rentable = Unidad.objects.create(
            numero_economico='ECO-R', placa='RRR-111', tipo='LOCAL', año=2020,
            capacidad_combustible=Decimal('200.00'), rendimiento_esperado=Decimal('3.00'),
        )
        self.unidad_perdida = Unidad.objects.create(
            numero_economico='ECO-P', placa='PPP-222', tipo='LOCAL', año=2020,
            capacidad_combustible=Decimal('200.00'), rendimiento_esperado=Decimal('3.00'),
        )
        self.operador = Operador.objects.create(nombre='Juan Pérez', tipo='LOCAL')
        self.despachador = Despachador.objects.create(nombre='Pedro López')

        aware = lambda y, m, d, h=8: timezone.make_aware(timezone.datetime(y, m, d, h))

        # Unidad rentable: ingreso alto, gasto bajo
        BitacoraViaje.objects.create(
            operador=self.operador, unidad=self.unidad_rentable, modalidad='LOCAL',
            fecha_carga=aware(2026, 6, 1), fecha_salida=aware(2026, 6, 1),
            fecha_llegada=aware(2026, 6, 2), destino='Destino rentable',
            ingreso_calculado=Decimal('5000.00'),
        )
        CargaCombustible.objects.create(
            despachador=self.despachador, unidad=self.unidad_rentable, cantidad_litros=Decimal('50.00'),
            kilometraje_actual=1000, nivel_combustible_inicial='MEDIO', estado_candado_anterior='NORMAL',
            fecha_hora_inicio=aware(2026, 6, 3), tipo_flujo='LOCAL', estado='COMPLETADO',
            costo_calculado=Decimal('500.00'),
        )

        # Unidad en pérdida: sin ingreso, solo gasto
        CargaCombustible.objects.create(
            despachador=self.despachador, unidad=self.unidad_perdida, cantidad_litros=Decimal('200.00'),
            kilometraje_actual=1000, nivel_combustible_inicial='MEDIO', estado_candado_anterior='NORMAL',
            fecha_hora_inicio=aware(2026, 6, 4), tipo_flujo='LOCAL', estado='COMPLETADO',
            costo_calculado=Decimal('3000.00'),
        )

        self.periodo_inicio = date(2026, 6, 1)
        self.periodo_fin = date(2026, 6, 30)

    def test_resumen_cuenta_unidades_en_utilidad_y_perdida(self):
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        self.assertEqual(datos['resumen']['total_unidades'], 2)
        self.assertEqual(datos['resumen']['unidades_en_utilidad'], 1)
        self.assertEqual(datos['resumen']['unidades_en_perdida'], 1)

    def test_resumen_incluye_totales_e_indicadores_de_cobertura(self):
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        resumen = datos['resumen']
        self.assertEqual(resumen['ingresos_totales'], 5000.0)
        self.assertEqual(resumen['gasto_total'], 3500.0)
        self.assertEqual(resumen['utilidad_total'], 1500.0)
        self.assertIn('bitacoras_excluidas', resumen)
        self.assertIn('cargas_excluidas', resumen)

    def test_filas_usa_numero_economico_y_valores_float(self):
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        fila_rentable = next(f for f in datos['filas'] if f['unidad'] == 'ECO-R')
        self.assertEqual(fila_rentable['ingresos'], 5000.0)
        self.assertIsInstance(fila_rentable['ingresos'], float)
        self.assertEqual(fila_rentable['utilidad'], 4500.0)

    def test_identifica_unidad_mas_rentable_y_de_mayor_perdida(self):
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        self.assertEqual(datos['unidad_mas_rentable']['unidad'], 'ECO-R')
        self.assertEqual(datos['unidad_mayor_perdida']['unidad'], 'ECO-P')

    def test_unidad_sin_actividad_no_cuenta_como_utilidad_ni_perdida(self):
        Unidad.objects.create(
            numero_economico='ECO-INACTIVA', placa='ZZZ-000', tipo='LOCAL', año=2020,
            capacidad_combustible=Decimal('200.00'), rendimiento_esperado=Decimal('3.00'),
        )
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        self.assertEqual(datos['resumen']['total_unidades'], 3)
        self.assertEqual(datos['resumen']['unidades_en_utilidad'], 1)
        self.assertEqual(datos['resumen']['unidades_en_perdida'], 1)
        self.assertEqual(datos['resumen']['unidades_sin_actividad'], 1)
        # La unidad inactiva no debe convertirse en "más rentable" ni "mayor pérdida"
        self.assertEqual(datos['unidad_mas_rentable']['unidad'], 'ECO-R')
        self.assertEqual(datos['unidad_mayor_perdida']['unidad'], 'ECO-P')

    def test_tipo_y_titulo_correctos(self):
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        self.assertEqual(datos['tipo'], 'UNIDADES_BALANZA_UTILIDAD')
        self.assertIn('Balanza de Utilidad', datos['titulo'])

    def test_sin_unidades_activas_no_lanza_error(self):
        Unidad.objects.all().update(activa=False)
        datos = generar_balanza_utilidad(self.periodo_inicio, self.periodo_fin)
        self.assertEqual(datos['filas'], [])
        self.assertIsNone(datos['unidad_mas_rentable'])
        self.assertIsNone(datos['unidad_mayor_perdida'])

    def test_generador_registrado_en_generadores(self):
        self.assertIn('UNIDADES_BALANZA_UTILIDAD', GENERADORES)
        self.assertIs(GENERADORES['UNIDADES_BALANZA_UTILIDAD'], generar_balanza_utilidad)


class PromptBalanzaUtilidadTests(TestCase):
    def test_prompt_incluye_unidades_extremas_y_totales(self):
        from modulos.reportes.generadores.narrativa import _prompt_unidades_balanza_utilidad

        resumen = {
            'total_unidades': 2, 'unidades_en_utilidad': 1, 'unidades_en_perdida': 1,
            'ingresos_totales': 5000.0, 'gasto_total': 3500.0, 'utilidad_total': 1500.0,
            'bitacoras_excluidas': 3, 'cargas_excluidas': 1,
        }
        datos = {
            'unidad_mas_rentable': {'unidad': 'ECO-R', 'utilidad': 4500.0},
            'unidad_mayor_perdida': {'unidad': 'ECO-P', 'utilidad': -3000.0},
        }
        prompt, max_tokens = _prompt_unidades_balanza_utilidad(
            resumen, datos, '2026-06-01', '2026-06-30'
        )
        self.assertIn('ECO-R', prompt)
        self.assertIn('ECO-P', prompt)
        self.assertIn('1,500.00', prompt)
        self.assertIn('3 bitácora(s) sin tarifa vigente', prompt)  # bitácoras excluidas mencionadas
        self.assertEqual(max_tokens, 500)

    def test_generar_narrativa_despacha_al_prompt_especializado(self):
        from unittest.mock import patch
        from modulos.reportes.generadores.narrativa import generar_narrativa

        resumen = {'total_unidades': 0, 'unidades_en_utilidad': 0, 'unidades_en_perdida': 0}
        datos = {'unidad_mas_rentable': None, 'unidad_mayor_perdida': None}

        with patch('modulos.reportes.generadores.narrativa.settings.IA_HABILITADA', True), \
             patch('config.services.claude_service.ClaudeService') as MockClaude:
            instancia = MockClaude.return_value
            instancia.completar.return_value = 'Narrativa de prueba'

            resultado = generar_narrativa(
                tipo_reporte='UNIDADES_BALANZA_UTILIDAD', resumen=resumen,
                periodo_inicio='2026-06-01', periodo_fin='2026-06-30', datos=datos,
            )

        self.assertEqual(resultado, 'Narrativa de prueba')
        args, kwargs = instancia.completar.call_args
        self.assertIn('Balanza', kwargs['prompt'] if 'prompt' in kwargs else args[0])
        self.assertEqual(kwargs['max_tokens'], 500)
        self.assertIn('salud financiera de la flotilla', kwargs['prompt'] if 'prompt' in kwargs else args[0])


class GenerarContenedoresPorOperadorTests(TestCase):
    def setUp(self):
        from modulos.modulacion.models import Agencia, TerminalPortuaria

        self.agencia = Agencia.objects.create(nombre='LOGINCO')
        self.terminal = TerminalPortuaria.objects.create(nombre='TIMSA')
        self.op_a = Operador.objects.create(nombre='Ana Local', tipo='LOCAL')
        self.op_b = Operador.objects.create(nombre='Beto Local', tipo='LOCAL')

    def _modulacion(self, operador=None, fecha_retiro=None, contenedor='ABCU1234567'):
        from modulos.modulacion.models import Modulacion

        m = Modulacion.objects.create(
            agencia=self.agencia,
            terminal_portuaria=self.terminal,
            tipo_contenedor='40HC',
            peso_toneladas=Decimal('18.00'),
            contenedor=contenedor,
            operador=operador,
        )
        if fecha_retiro is not None:
            m.fecha_retiro = fecha_retiro
            m.save(update_fields=['fecha_retiro'])
        return m

    def test_cuenta_por_operador_dentro_del_rango(self):
        from modulos.reportes.generadores.modulacion import generar_contenedores_por_operador

        base = timezone.make_aware(timezone.datetime(2026, 8, 25, 10, 0))  # lunes W35
        self._modulacion(self.op_a, base, 'AAAA0000001')
        self._modulacion(self.op_a, base + timedelta(days=1), 'AAAA0000002')
        self._modulacion(self.op_b, base + timedelta(days=2), 'BBBB0000001')

        datos = generar_contenedores_por_operador(date(2026, 8, 24), date(2026, 8, 30))

        self.assertEqual(datos['resumen']['total_contenedores'], 3)
        self.assertEqual(datos['resumen']['operadores_activos'], 2)
        self.assertEqual(datos['resumen']['operador_top'], 'Ana Local')
        self.assertEqual(datos['resumen']['contenedores_operador_top'], 2)
        totales = {f['operador']: f['contenedores'] for f in datos['totales_operador']}
        self.assertEqual(totales, {'Ana Local': 2, 'Beto Local': 1})

    def test_excluye_sin_operador_y_fuera_de_rango(self):
        from modulos.reportes.generadores.modulacion import generar_contenedores_por_operador

        dentro = timezone.make_aware(timezone.datetime(2026, 8, 26, 9, 0))
        fuera = timezone.make_aware(timezone.datetime(2026, 7, 1, 9, 0))
        self._modulacion(self.op_a, dentro, 'AAAA0000010')
        self._modulacion(None, dentro, 'CCCC0000001')          # sin operador
        self._modulacion(self.op_b, fuera, 'BBBB0000010')        # fuera de rango
        self._modulacion(self.op_b, None, 'BBBB0000011')         # sin fecha_retiro

        datos = generar_contenedores_por_operador(date(2026, 8, 24), date(2026, 8, 30))
        self.assertEqual(datos['resumen']['total_contenedores'], 1)
        self.assertEqual(datos['resumen']['operadores_activos'], 1)

    def test_agrupa_por_semana_iso(self):
        from modulos.reportes.generadores.modulacion import generar_contenedores_por_operador

        w35 = timezone.make_aware(timezone.datetime(2026, 8, 25, 8, 0))
        w36 = timezone.make_aware(timezone.datetime(2026, 9, 1, 8, 0))
        self._modulacion(self.op_a, w35, 'AAAA0000020')
        self._modulacion(self.op_a, w36, 'AAAA0000021')

        datos = generar_contenedores_por_operador(date(2026, 8, 24), date(2026, 9, 6))
        semanas = sorted({f['semana'] for f in datos['filas']})
        self.assertEqual(len(semanas), 2)
        self.assertTrue(any('W35' in s for s in semanas))
        self.assertTrue(any('W36' in s for s in semanas))

    def test_registrado_en_generadores_del_comando(self):
        from modulos.reportes.management.commands.generar_reportes import GENERADORES as CMD_GEN

        self.assertIn('MODULACION_CONTENEDORES_OPERADOR', CMD_GEN)


class ContenedoresPorOperadorViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='rep-tester', password='pass12345')
        self.client.login(username='rep-tester', password='pass12345')
        self.url = '/reportes/modulacion/contenedores-por-operador/'

    def test_responde_200_con_rango_default(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('datos', response.context)
        self.assertIn('desde', response.context)
        self.assertIn('hasta', response.context)

    def test_rango_explicito_se_respeta(self):
        response = self.client.get(self.url, {'desde': '2026-08-01', 'hasta': '2026-08-31'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(str(response.context['desde']), '2026-08-01')
        self.assertEqual(str(response.context['hasta']), '2026-08-31')

    def test_rango_invalido_cae_a_default(self):
        response = self.client.get(self.url, {'desde': 'no-es-fecha', 'hasta': ''})
        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(response.context['desde'], response.context['hasta'])

    def test_requiere_login(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
