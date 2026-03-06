"""
Management command: generar_reportes

Ejecuta los reportes programados cuya fecha de ejecución ha llegado.
Diseñado para correr vía cron (GitHub Actions, DigitalOcean Scheduler, etc.):

    python manage.py generar_reportes

Flags opcionales:
    --forzar-id <id>   Ejecuta un reporte específico sin importar si es_debido()
    --dry-run          Muestra qué reportes se ejecutarían sin enviar correos
"""

import logging
from datetime import date, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color

from django.core.management.base import BaseCommand
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils import timezone
from django.conf import settings

from modulos.reportes.models import ConfiguracionReporte, ReporteGenerado
from modulos.reportes.generadores import almacen as gen_almacen
from modulos.reportes.generadores import combustible as gen_combustible

logger = logging.getLogger(__name__)

GENERADORES = {
    **gen_almacen.GENERADORES,
    **gen_combustible.GENERADORES,
}


def _periodo(frecuencia: str, dia_semana, dia_mes):
    """Calcula (periodo_inicio, periodo_fin) según la frecuencia del reporte."""
    hoy = timezone.now().date()
    if frecuencia == 'DIARIO':
        inicio = hoy - timedelta(days=1)
        fin = hoy - timedelta(days=1)
    elif frecuencia == 'SEMANAL':
        fin = hoy - timedelta(days=1)
        inicio = fin - timedelta(days=6)
    else:  # MENSUAL
        fin = hoy - timedelta(days=1)
        inicio = date(hoy.year, hoy.month, 1) - timedelta(days=1)
        inicio = date(inicio.year, inicio.month, 1)
    return inicio, fin


def _generar_excel(datos: dict) -> bytes:
    """Genera un archivo Excel con el detalle del período reportado."""
    wb = openpyxl.Workbook()

    # --- Hoja de detalle ---
    ws = wb.active
    titulo_hoja = datos.get('titulo', 'Reporte').replace('/', '-').replace('\\', '-')[:31]
    ws.title = titulo_hoja

    filas = datos.get('filas', [])
    if filas:
        headers = list(filas[0].keys())
        fill_azul = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        font_blanco_bold = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = font_blanco_bold
            cell.fill = fill_azul
            cell.alignment = Alignment(horizontal='center')

        # Detectar qué columnas son de foto (contienen URLs)
        foto_headers = {h for h in headers if h.startswith('foto_')}
        font_link = Font(color='1D4ED8', underline='single')

        for row_idx, fila in enumerate(filas, 2):
            for col_idx, header in enumerate(headers, 1):
                valor = fila.get(header) or ''
                cell = ws.cell(row=row_idx, column=col_idx)
                if header in foto_headers and valor:
                    cell.value = 'Ver foto'
                    cell.hyperlink = valor
                    cell.font = font_link
                else:
                    cell.value = valor if valor != '' else None

        # Ajustar ancho de columnas (fotos fijas en 12)
        for col in ws.columns:
            header_cell = col[0]
            letter = header_cell.column_letter
            if header_cell.value and str(header_cell.value).lower().startswith('foto'):
                ws.column_dimensions[letter].width = 12
            else:
                max_len = max((len(str(cell.value or '')) for cell in col), default=8)
                ws.column_dimensions[letter].width = min(max_len + 4, 40)

    # --- Hoja de resumen ---
    ws_res = wb.create_sheet('Resumen')
    ws_res['A1'] = 'Métrica'
    ws_res['B1'] = 'Valor'
    ws_res['A1'].font = Font(bold=True)
    ws_res['B1'].font = Font(bold=True)
    for idx, (k, v) in enumerate(datos.get('resumen', {}).items(), 2):
        ws_res.cell(row=idx, column=1, value=k.replace('_', ' ').title())
        ws_res.cell(row=idx, column=2, value=v)
    ws_res.column_dimensions['A'].width = 30
    ws_res.column_dimensions['B'].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _enviar_email(config: ConfiguracionReporte, datos: dict, dry_run: bool,
                  excel_bytes: bytes = None) -> list:
    """Envía el reporte por correo. Devuelve lista de destinatarios enviados."""
    destinatarios = config.get_destinatarios_list()
    if not destinatarios:
        logger.warning("Reporte %s sin destinatarios", config.nombre)
        return []

    if dry_run:
        logger.info("[DRY-RUN] Enviaría a: %s", ", ".join(destinatarios))
        return destinatarios

    asunto = f"[BitacoraKasu] {datos['titulo']}"
    texto_plano = f"{datos['titulo']}\n\nGenerado: {datos['generado_en']}\n\nResumen:\n"
    for k, v in datos.get('resumen', {}).items():
        texto_plano += f"  {k}: {v}\n"

    html = render_to_string('reportes/email/reporte_base.html', {'datos': datos, 'config': config})

    msg = EmailMultiAlternatives(
        subject=asunto,
        body=texto_plano,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=destinatarios,
    )
    msg.attach_alternative(html, 'text/html')

    if excel_bytes:
        nombre_excel = (
            f"reporte_{datos.get('tipo', 'reporte')}_"
            f"{datos['periodo_inicio']}_{datos['periodo_fin']}.xlsx"
        )
        msg.attach(
            nombre_excel,
            excel_bytes,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    msg.send(fail_silently=False)
    return destinatarios


class Command(BaseCommand):
    help = 'Ejecuta los reportes programados pendientes y los envía por correo.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--forzar-id',
            type=int,
            metavar='ID',
            help='Ejecuta el reporte con este ID sin verificar si es_debido()',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Muestra qué se ejecutaría sin enviar correos ni guardar historial',
        )

    def handle(self, *args, **options):
        forzar_id = options.get('forzar_id')
        dry_run = options.get('dry_run', False)

        if forzar_id:
            configs = ConfiguracionReporte.objects.filter(id=forzar_id, activo=True)
        else:
            configs = ConfiguracionReporte.objects.filter(activo=True)

        ejecutados = 0
        errores = 0

        for config in configs:
            if not forzar_id and not config.es_debido():
                self.stdout.write(f"  SKIP  {config.nombre} (no es su momento)")
                continue

            generador = GENERADORES.get(config.tipo_reporte)
            if not generador:
                self.stderr.write(f"  ERROR {config.nombre}: sin generador para '{config.tipo_reporte}'")
                errores += 1
                continue

            periodo_inicio, periodo_fin = _periodo(config.frecuencia, config.dia_semana, config.dia_mes)

            self.stdout.write(f"  RUN   {config.nombre} ({periodo_inicio} → {periodo_fin})")

            try:
                datos = generador(periodo_inicio, periodo_fin)
                excel_bytes = _generar_excel(datos) if config.adjuntar_excel else None
                enviados = _enviar_email(config, datos, dry_run, excel_bytes=excel_bytes)

                if not dry_run:
                    ReporteGenerado.objects.create(
                        configuracion=config,
                        periodo_inicio=periodo_inicio,
                        periodo_fin=periodo_fin,
                        estado='GENERADO',
                        destinatarios_enviados=', '.join(enviados),
                        resumen=datos.get('resumen', {}),
                    )
                    config.ultimo_envio = timezone.now()
                    config.save(update_fields=['ultimo_envio'])

                ejecutados += 1
                self.stdout.write(self.style.SUCCESS(f"  OK    {config.nombre}"))

            except Exception as exc:
                logger.exception("Error generando reporte %s", config.nombre)
                errores += 1
                if not dry_run:
                    ReporteGenerado.objects.create(
                        configuracion=config,
                        periodo_inicio=periodo_inicio,
                        periodo_fin=periodo_fin,
                        estado='ERROR',
                        mensaje_error=str(exc),
                    )
                self.stderr.write(self.style.ERROR(f"  FAIL  {config.nombre}: {exc}"))

        self.stdout.write(f"\nEjecutados: {ejecutados}  Errores: {errores}")
