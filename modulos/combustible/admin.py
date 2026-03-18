from django.contrib import admin
from django.utils.html import format_html
from django.urls import path
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count, Max, Min, Q
from django.db.models.functions import TruncDay, TruncMonth
from datetime import date, timedelta
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Despachador, CargaCombustible, FotoCandadoNuevo, AlertaCombustible


@admin.register(Despachador)
class DespachadorAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'telefono', 'activo', 'total_cargas', 'created_at']
    list_filter = ['activo', 'created_at']
    search_fields = ['nombre', 'telefono']
    ordering = ['nombre']

    fieldsets = (
        ('Información Personal', {
            'fields': ('nombre', 'telefono')
        }),
        ('Estado', {
            'fields': ('activo',)
        }),
    )

    def total_cargas(self, obj):
        count = obj.cargas.count()
        return format_html(
            '<span style="background: #3b82f6; color: white; padding: 2px 8px; border-radius: 4px;">{}</span>',
            count
        )
    total_cargas.short_description = 'Total cargas'


class FotoCandadoNuevoInline(admin.TabularInline):
    model = FotoCandadoNuevo
    extra = 1
    fields = ['foto', 'descripcion']
    readonly_fields = []


@admin.register(CargaCombustible)
class CargaCombustibleAdmin(admin.ModelAdmin):
    inlines = [FotoCandadoNuevoInline]
    list_display = [
        'id', 'unidad_info', 'despachador', 'cantidad_litros',
        'estado_badge', 'alerta_badge', 'fecha_hora_inicio', 'tiempo_carga'
    ]
    list_filter = [
        'estado', 'estado_candado_anterior', 'nivel_combustible_inicial',
        'fecha_hora_inicio', 'unidad__tipo'
    ]
    search_fields = [
        'unidad__numero_economico', 'unidad__placa',
        'despachador__nombre', 'notas'
    ]
    readonly_fields = [
        'tiempo_carga_minutos', 'created_at', 'updated_at',
        'preview_fotos'
    ]
    date_hierarchy = 'fecha_hora_inicio'
    ordering = ['-fecha_hora_inicio']

    fieldsets = (
        ('Información Principal', {
            'fields': ('despachador', 'unidad', 'estado')
        }),
        ('Datos de Carga', {
            'fields': ('cantidad_litros', 'fecha_hora_inicio', 'fecha_hora_fin', 'tiempo_carga_minutos')
        }),
        ('Datos del Tablero', {
            'fields': ('kilometraje_actual', 'nivel_combustible_inicial')
        }),
        ('Estado del Candado', {
            'fields': ('estado_candado_anterior', 'observaciones_candado')
        }),
        ('Fotografías', {
            'fields': (
                'foto_numero_economico', 'foto_tablero', 'foto_candado_anterior',
                'foto_candado_nuevo', 'foto_ticket', 'preview_fotos'
            ),
            'classes': ('collapse',)
        }),
        ('Notas', {
            'fields': ('notas',)
        }),
        ('Metadatos', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    # ─── Display methods ───────────────────────────────────────────────────────

    def unidad_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small style="color: #666;">{}</small>',
            obj.unidad.numero_economico,
            obj.unidad.placa
        )
    unidad_info.short_description = 'Unidad'

    def estado_badge(self, obj):
        colors = {
            'INICIADO': '#6b7280',
            'EN_PROCESO': '#8b5cf6',
            'COMPLETADO': '#10b981',
            'CANCELADO': '#ef4444',
        }
        color = colors.get(obj.estado, '#6b7280')
        return format_html(
            '<span style="background: {}; color: white; padding: 4px 8px; border-radius: 4px; font-size: 11px; font-weight: bold;">{}</span>',
            color,
            obj.get_estado_display()
        )
    estado_badge.short_description = 'Estado'

    def alerta_badge(self, obj):
        if obj.tiene_alertas():
            return format_html(
                '<span style="background: #ef4444; color: white; padding: 4px 8px; border-radius: 4px; font-size: 11px;">⚠️ {}</span>',
                obj.get_estado_candado_anterior_display()
            )
        return format_html('<span style="color: #10b981; font-weight: bold;">✓ Normal</span>')
    alerta_badge.short_description = 'Candado'

    def tiempo_carga(self, obj):
        if obj.tiempo_carga_minutos:
            return format_html(
                '<span style="background: #8b5cf6; color: white; padding: 2px 8px; border-radius: 4px;">{} min</span>',
                obj.tiempo_carga_minutos
            )
        return '-'
    tiempo_carga.short_description = 'Tiempo'

    def preview_fotos(self, obj):
        html = '<div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 10px;">'
        fotos = [
            ('Número Económico', obj.foto_numero_economico),
            ('Tablero', obj.foto_tablero),
            ('Candado Anterior', obj.foto_candado_anterior),
        ]
        fotos_candado = obj.fotos_candado_nuevo.all()
        if fotos_candado.exists():
            for foto_obj in fotos_candado:
                fotos.append((foto_obj.descripcion or 'Candado Nuevo', foto_obj.foto))
        elif obj.foto_candado_nuevo:
            fotos.append(('Candado Nuevo', obj.foto_candado_nuevo))
        fotos.append(('Ticket', obj.foto_ticket))

        for nombre, foto in fotos:
            if foto:
                html += f'''
                <div style="border: 1px solid #e5e7eb; border-radius: 8px; padding: 8px;">
                    <p style="font-weight: bold; font-size: 12px; margin-bottom: 5px;">{nombre}</p>
                    <a href="{foto.url}" target="_blank">
                        <img src="{foto.url}" style="width: 100%; height: 120px; object-fit: cover; border-radius: 4px;">
                    </a>
                </div>
                '''
        html += '</div>'
        return format_html(html)
    preview_fotos.short_description = 'Vista previa de fotografías'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('despachador', 'unidad')

    actions = ['marcar_completado', 'exportar_reporte']

    def marcar_completado(self, request, queryset):
        updated = queryset.update(estado='COMPLETADO')
        self.message_user(request, f'{updated} cargas marcadas como completadas.')
    marcar_completado.short_description = 'Marcar como completadas'

    def exportar_reporte(self, request, queryset):
        self.message_user(request, 'Función de exportación en desarrollo.')
    exportar_reporte.short_description = 'Exportar reporte'

    # ─── URLs personalizadas ───────────────────────────────────────────────────

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'estadisticas/',
                self.admin_site.admin_view(self.estadisticas_view),
                name='combustible_cargacombustible_estadisticas',
            ),
            path(
                'estadisticas/exportar-excel/',
                self.admin_site.admin_view(self.exportar_excel_view),
                name='combustible_cargacombustible_estadisticas_excel',
            ),
        ]
        return custom_urls + urls

    # ─── Vista de estadísticas ─────────────────────────────────────────────────

    def estadisticas_view(self, request):
        hoy = date.today()

        desde_str = request.GET.get('desde', (hoy - timedelta(days=30)).isoformat())
        hasta_str = request.GET.get('hasta', hoy.isoformat())
        try:
            desde = date.fromisoformat(desde_str)
        except (ValueError, TypeError):
            desde = hoy - timedelta(days=30)
        try:
            hasta = date.fromisoformat(hasta_str)
        except (ValueError, TypeError):
            hasta = hoy

        qs = CargaCombustible.objects.filter(
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=desde,
            fecha_hora_inicio__date__lte=hasta,
        ).select_related('unidad', 'despachador')

        qs_todos = CargaCombustible.objects.filter(
            fecha_hora_inicio__date__gte=desde,
            fecha_hora_inicio__date__lte=hasta,
        )

        # ─── KPIs ─────────────────────────────────────────────────────────────
        totales = qs.aggregate(
            total_cargas=Count('id'),
            total_litros=Sum('cantidad_litros'),
            promedio_litros=Avg('cantidad_litros'),
            promedio_tiempo=Avg('tiempo_carga_minutos'),
            max_litros=Max('cantidad_litros'),
            min_litros=Min('cantidad_litros'),
        )
        for key in ('total_litros', 'promedio_litros', 'promedio_tiempo', 'max_litros', 'min_litros'):
            totales[key] = round(float(totales[key] or 0), 2)

        estados_count = {
            'COMPLETADO': qs_todos.filter(estado='COMPLETADO').count(),
            'CANCELADO': qs_todos.filter(estado='CANCELADO').count(),
            'EN_PROCESO': qs_todos.filter(estado='EN_PROCESO').count(),
            'INICIADO': qs_todos.filter(estado='INICIADO').count(),
        }
        total_alertas_pendientes = AlertaCombustible.objects.filter(
            resuelta=False,
            carga__fecha_hora_inicio__date__gte=desde,
            carga__fecha_hora_inicio__date__lte=hasta,
        ).count()

        # ─── Nivel inicial ─────────────────────────────────────────────────────
        NIVEL_LABELS = {'VACIO': 'Vacío', 'CUARTO': '1/4', 'MEDIO': '1/2', 'TRES_CUARTOS': '3/4'}
        NIVEL_ORDEN = ['VACIO', 'CUARTO', 'MEDIO', 'TRES_CUARTOS']

        niveles_raw = {
            n['nivel_combustible_inicial']: n
            for n in qs.values('nivel_combustible_inicial').annotate(
                total=Count('id'),
                litros=Sum('cantidad_litros'),
                promedio=Avg('cantidad_litros'),
            )
        }
        niveles = []
        for codigo in NIVEL_ORDEN:
            n = niveles_raw.get(codigo, {'total': 0, 'litros': 0, 'promedio': 0})
            niveles.append({
                'codigo': codigo,
                'label': NIVEL_LABELS[codigo],
                'total': n['total'],
                'litros': round(float(n['litros'] or 0), 2),
                'promedio': round(float(n['promedio'] or 0), 2),
            })

        nivel_labels_json = json.dumps([n['label'] for n in niveles])
        nivel_totales_json = json.dumps([n['total'] for n in niveles])
        nivel_promedios_json = json.dumps([n['promedio'] for n in niveles])
        nivel_litros_json = json.dumps([n['litros'] for n in niveles])

        # ─── Cargas por día ────────────────────────────────────────────────────
        cargas_dia = list(
            qs.annotate(dia=TruncDay('fecha_hora_inicio'))
            .values('dia')
            .annotate(total=Count('id'), litros=Sum('cantidad_litros'))
            .order_by('dia')
        )
        dias_labels_json = json.dumps([c['dia'].strftime('%d/%m') for c in cargas_dia])
        dias_litros_json = json.dumps([round(float(c['litros']), 2) for c in cargas_dia])
        dias_cargas_json = json.dumps([c['total'] for c in cargas_dia])

        # ─── Top 10 unidades ───────────────────────────────────────────────────
        top_unidades = list(
            qs.values('unidad__numero_economico', 'unidad__placa', 'unidad__tipo')
            .annotate(
                total_cargas=Count('id'),
                total_litros=Sum('cantidad_litros'),
                promedio_litros=Avg('cantidad_litros'),
            )
            .order_by('-total_litros')[:10]
        )
        for u in top_unidades:
            u['total_litros'] = round(float(u['total_litros'] or 0), 2)
            u['promedio_litros'] = round(float(u['promedio_litros'] or 0), 2)

        top_unidades_labels_json = json.dumps([u['unidad__numero_economico'] for u in top_unidades])
        top_unidades_litros_json = json.dumps([u['total_litros'] for u in top_unidades])
        top_unidades_cargas_json = json.dumps([u['total_cargas'] for u in top_unidades])

        # ─── Estado de candados ────────────────────────────────────────────────
        CANDADO_LABELS = {
            'NORMAL': 'Normal', 'ALTERADO': 'Alterado',
            'VIOLADO': 'Violado', 'SIN_CANDADO': 'Sin Candado',
        }
        candados_raw = qs.values('estado_candado_anterior').annotate(total=Count('id')).order_by('-total')
        candados = [
            {'label': CANDADO_LABELS.get(c['estado_candado_anterior'], c['estado_candado_anterior']), 'total': c['total']}
            for c in candados_raw
        ]
        candados_labels_json = json.dumps([c['label'] for c in candados])
        candados_totales_json = json.dumps([c['total'] for c in candados])

        # ─── Por despachador ───────────────────────────────────────────────────
        por_desp = list(
            qs.values('despachador__nombre')
            .annotate(
                total_cargas=Count('id'),
                total_litros=Sum('cantidad_litros'),
                promedio_litros=Avg('cantidad_litros'),
            )
            .order_by('-total_litros')
        )
        for d in por_desp:
            d['total_litros'] = round(float(d['total_litros'] or 0), 2)
            d['promedio_litros'] = round(float(d['promedio_litros'] or 0), 2)

        desp_labels_json = json.dumps([d['despachador__nombre'] for d in por_desp])
        desp_litros_json = json.dumps([d['total_litros'] for d in por_desp])
        desp_cargas_json = json.dumps([d['total_cargas'] for d in por_desp])

        # ─── Historial mensual últimos 12 meses ────────────────────────────────
        mes = hoy.month - 11
        anio = hoy.year
        if mes <= 0:
            mes += 12
            anio -= 1
        hist_desde = date(anio, mes, 1)

        hist_mensual = list(
            CargaCombustible.objects.filter(estado='COMPLETADO', fecha_hora_inicio__date__gte=hist_desde)
            .annotate(mes=TruncMonth('fecha_hora_inicio'))
            .values('mes')
            .annotate(total_cargas=Count('id'), total_litros=Sum('cantidad_litros'))
            .order_by('mes')
        )
        for m in hist_mensual:
            m['total_litros'] = round(float(m['total_litros']), 2)

        hist_labels_json = json.dumps([m['mes'].strftime('%b %Y') for m in hist_mensual])
        hist_litros_json = json.dumps([m['total_litros'] for m in hist_mensual])
        hist_cargas_json = json.dumps([m['total_cargas'] for m in hist_mensual])

        # ─── Alertas por tipo ──────────────────────────────────────────────────
        ALERTA_LABELS = dict(AlertaCombustible.TIPO_CHOICES)
        alertas_data = list(
            AlertaCombustible.objects.filter(
                carga__fecha_hora_inicio__date__gte=desde,
                carga__fecha_hora_inicio__date__lte=hasta,
            )
            .values('tipo_alerta')
            .annotate(
                total=Count('id'),
                pendientes=Count('id', filter=Q(resuelta=False)),
            )
            .order_by('-total')
        )
        alertas = [
            {
                'label': ALERTA_LABELS.get(a['tipo_alerta'], a['tipo_alerta']),
                'total': a['total'],
                'pendientes': a['pendientes'],
                'resueltas': a['total'] - a['pendientes'],
            }
            for a in alertas_data
        ]
        alertas_labels_json = json.dumps([a['label'] for a in alertas])
        alertas_totales_json = json.dumps([a['total'] for a in alertas])
        alertas_pendientes_json = json.dumps([a['pendientes'] for a in alertas])

        # ─── Proyecciones ──────────────────────────────────────────────────────
        dias_en_periodo = max((hasta - desde).days + 1, 1)
        total_litros_float = totales['total_litros']
        total_cargas_int = totales['total_cargas'] or 0

        promedio_diario_litros = total_litros_float / dias_en_periodo
        promedio_diario_cargas = total_cargas_int / dias_en_periodo

        proyeccion_30_dias = round(promedio_diario_litros * 30, 2)
        proyeccion_90_dias = round(promedio_diario_litros * 90, 2)
        proyeccion_anual   = round(promedio_diario_litros * 365, 2)
        proyeccion_cargas_30 = round(promedio_diario_cargas * 30)

        # Tendencia intra-período: primera mitad vs segunda mitad
        mitad = desde + timedelta(days=dias_en_periodo // 2)
        litros_primera = float(
            qs.filter(fecha_hora_inicio__date__lt=mitad)
            .aggregate(t=Sum('cantidad_litros'))['t'] or 0
        )
        litros_segunda = float(
            qs.filter(fecha_hora_inicio__date__gte=mitad)
            .aggregate(t=Sum('cantidad_litros'))['t'] or 0
        )
        if litros_primera > 0:
            variacion_interna = round(((litros_segunda - litros_primera) / litros_primera) * 100, 1)
        else:
            variacion_interna = 0
        if variacion_interna > 5:
            tendencia_label = 'alza'
            tendencia_color = 'rojo'
        elif variacion_interna < -5:
            tendencia_label = 'baja'
            tendencia_color = 'verde'
        else:
            tendencia_label = 'estable'
            tendencia_color = 'azul'

        # Comparación con período anterior equivalente
        per_ant_hasta = desde - timedelta(days=1)
        per_ant_desde = per_ant_hasta - timedelta(days=dias_en_periodo - 1)
        litros_periodo_anterior = float(
            CargaCombustible.objects.filter(
                estado='COMPLETADO',
                fecha_hora_inicio__date__gte=per_ant_desde,
                fecha_hora_inicio__date__lte=per_ant_hasta,
            ).aggregate(t=Sum('cantidad_litros'))['t'] or 0
        )
        cargas_periodo_anterior = CargaCombustible.objects.filter(
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=per_ant_desde,
            fecha_hora_inicio__date__lte=per_ant_hasta,
        ).count()

        if litros_periodo_anterior > 0:
            variacion_vs_anterior = round(
                ((total_litros_float - litros_periodo_anterior) / litros_periodo_anterior) * 100, 1
            )
        else:
            variacion_vs_anterior = None

        # ─── Interpretaciones automáticas ─────────────────────────────────────
        interpretaciones = []

        # 1. Nivel inicial dominante
        nivel_dominante = max(niveles, key=lambda n: n['total']) if niveles else None
        if nivel_dominante and nivel_dominante['total'] > 0:
            pct_dominante = round(nivel_dominante['total'] / total_cargas_int * 100) if total_cargas_int else 0
            if nivel_dominante['codigo'] in ('VACIO', 'CUARTO'):
                interpretaciones.append({
                    'semaforo': 'amarillo',
                    'titulo': 'Nivel inicial bajo predominante',
                    'texto': (
                        f"El {pct_dominante}% de las cargas se realiza con el tanque en nivel "
                        f"'{nivel_dominante['label']}'. Las unidades llegan con poco combustible, "
                        f"lo que indica ciclos de consumo total entre cargas — patrón esperado en flotas de largo recorrido."
                    ),
                })
            else:
                interpretaciones.append({
                    'semaforo': 'verde',
                    'titulo': 'Nivel inicial adecuado',
                    'texto': (
                        f"El {pct_dominante}% de las cargas se efectúa con el tanque en '{nivel_dominante['label']}', "
                        f"lo que sugiere recargas preventivas antes de agotar el combustible. "
                        f"Buen indicador de planeación operativa."
                    ),
                })

        # 2. Correlación nivel inicial vs litros cargados
        nivel_vacio  = next((n for n in niveles if n['codigo'] == 'VACIO'), None)
        nivel_medio  = next((n for n in niveles if n['codigo'] == 'MEDIO'), None)
        if nivel_vacio and nivel_medio and nivel_vacio['promedio'] > 0 and nivel_medio['promedio'] > 0:
            diferencia_pct = round(
                ((nivel_vacio['promedio'] - nivel_medio['promedio']) / nivel_medio['promedio']) * 100, 1
            )
            if diferencia_pct > 10:
                interpretaciones.append({
                    'semaforo': 'verde',
                    'titulo': 'Correlación nivel-litros confirmada',
                    'texto': (
                        f"Las unidades que llegan con tanque vacío cargan en promedio {nivel_vacio['promedio']:.1f} lt "
                        f"({diferencia_pct}% más que las que llegan a la mitad con {nivel_medio['promedio']:.1f} lt). "
                        f"El sistema refleja fielmente el consumo real por ciclo de viaje."
                    ),
                })

        # 3. Tasa de anomalías en candados
        cargas_con_anomalia = qs.filter(
            estado_candado_anterior__in=['ALTERADO', 'VIOLADO', 'SIN_CANDADO']
        ).count()
        if total_cargas_int > 0:
            tasa_anomalia = round(cargas_con_anomalia / total_cargas_int * 100, 1)
            if tasa_anomalia == 0:
                interpretaciones.append({
                    'semaforo': 'verde',
                    'titulo': 'Sin anomalías en candados',
                    'texto': (
                        "El 100% de las cargas registró candados en estado NORMAL en el período. "
                        "No se detectaron indicios de manipulación indebida."
                    ),
                })
            elif tasa_anomalia <= 5:
                interpretaciones.append({
                    'semaforo': 'amarillo',
                    'titulo': f'Tasa de anomalías baja ({tasa_anomalia}%)',
                    'texto': (
                        f"{cargas_con_anomalia} de {total_cargas_int} cargas presentaron candado en estado anómalo "
                        f"({tasa_anomalia}%). Nivel manejable, pero se recomienda revisar cada caso."
                    ),
                })
            else:
                interpretaciones.append({
                    'semaforo': 'rojo',
                    'titulo': f'Alta tasa de anomalías en candados ({tasa_anomalia}%)',
                    'texto': (
                        f"{cargas_con_anomalia} de {total_cargas_int} cargas registraron problemas en el candado "
                        f"({tasa_anomalia}%). Se requiere revisión urgente del protocolo de seguridad."
                    ),
                })

        # 4. Tendencia de consumo en el período
        if total_cargas_int >= 4:
            if tendencia_label == 'alza':
                interpretaciones.append({
                    'semaforo': 'amarillo',
                    'titulo': f'Consumo en alza (+{variacion_interna}%)',
                    'texto': (
                        f"La segunda mitad del período muestra un incremento de {variacion_interna}% en litros "
                        f"respecto a la primera mitad. Podría indicar mayor actividad de la flota o una caída "
                        f"en el rendimiento de alguna(s) unidad(es)."
                    ),
                })
            elif tendencia_label == 'baja':
                interpretaciones.append({
                    'semaforo': 'verde',
                    'titulo': f'Consumo en descenso ({variacion_interna}%)',
                    'texto': (
                        f"La segunda mitad del período muestra una reducción de {abs(variacion_interna)}% en litros "
                        f"respecto a la primera mitad. Posible mejora en eficiencia o reducción de viajes."
                    ),
                })
            else:
                interpretaciones.append({
                    'semaforo': 'verde',
                    'titulo': 'Consumo estable en el período',
                    'texto': (
                        f"La variación entre ambas mitades del período es de solo {variacion_interna}%, "
                        f"lo que indica un patrón de consumo uniforme y predecible."
                    ),
                })

        # 5. Comparación vs período anterior
        if variacion_vs_anterior is not None:
            signo = '+' if variacion_vs_anterior >= 0 else ''
            if abs(variacion_vs_anterior) <= 10:
                sem = 'verde'
                lectura = 'consumo similar al período previo'
            elif variacion_vs_anterior > 10:
                sem = 'amarillo'
                lectura = 'consumo superior al período previo'
            else:
                sem = 'verde'
                lectura = 'consumo inferior al período previo'
            interpretaciones.append({
                'semaforo': sem,
                'titulo': f'vs período anterior: {signo}{variacion_vs_anterior}%',
                'texto': (
                    f"En el período equivalente anterior ({per_ant_desde.strftime('%d/%m/%Y')} – "
                    f"{per_ant_hasta.strftime('%d/%m/%Y')}) se cargaron "
                    f"{litros_periodo_anterior:,.1f} lt en {cargas_periodo_anterior} cargas. "
                    f"El período actual registra {signo}{variacion_vs_anterior}% ({lectura})."
                ),
            })

        # 6. Tasa de completación
        total_todas = qs_todos.count()
        if total_todas > 0:
            tasa_completacion = round(total_cargas_int / total_todas * 100, 1)
            canceladas = estados_count['CANCELADO']
            if tasa_completacion >= 90:
                interpretaciones.append({
                    'semaforo': 'verde',
                    'titulo': f'Alta tasa de completación ({tasa_completacion}%)',
                    'texto': (
                        f"{total_cargas_int} de {total_todas} cargas iniciadas fueron completadas exitosamente "
                        f"({tasa_completacion}%). Solo {canceladas} canceladas."
                    ),
                })
            else:
                interpretaciones.append({
                    'semaforo': 'amarillo',
                    'titulo': f'Tasa de completación moderada ({tasa_completacion}%)',
                    'texto': (
                        f"{canceladas} cargas fueron canceladas de {total_todas} iniciadas. "
                        f"Se recomienda identificar las causas de cancelación para reducirlas."
                    ),
                })

        # 7. Unidad con mayor consumo
        if top_unidades:
            top = top_unidades[0]
            interpretaciones.append({
                'semaforo': 'azul',
                'titulo': f'Mayor consumidor: Unidad {top["unidad__numero_economico"]}',
                'texto': (
                    f'La unidad {top["unidad__numero_economico"]} ({top["unidad__placa"]}) encabeza el consumo '
                    f'con {top["total_litros"]:,.2f} lt en {top["total_cargas"]} cargas '
                    f'(promedio {top["promedio_litros"]:,.2f} lt/carga). '
                    f'Representa el {round(top["total_litros"] / total_litros_float * 100, 1) if total_litros_float else 0}% '
                    f'del total de litros del período.'
                ),
            })

        # ─── Datos de proyección para gráfica ─────────────────────────────────
        # Proyección lineal simple para los próximos 6 meses usando promedio mensual
        # Se construye con los datos del historial mensual
        if len(hist_mensual) >= 2:
            ultimos_meses_litros = [m['total_litros'] for m in hist_mensual[-3:]]
            promedio_mensual_reciente = sum(ultimos_meses_litros) / len(ultimos_meses_litros)
        elif hist_mensual:
            promedio_mensual_reciente = hist_mensual[-1]['total_litros']
        else:
            promedio_mensual_reciente = promedio_diario_litros * 30

        # Etiquetas para próximos 6 meses
        MESES_ES = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                    'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        proy_labels = []
        proy_valores = []
        for i in range(1, 7):
            m_num = (hoy.month + i - 1) % 12 + 1
            a_num = hoy.year + (hoy.month + i - 1) // 12
            proy_labels.append(f'{MESES_ES[m_num - 1]} {a_num}')
            proy_valores.append(round(promedio_mensual_reciente, 2))

        hist_proy_labels = json.dumps(
            [m['mes'].strftime('%b %Y') for m in hist_mensual] + proy_labels
        )
        hist_proy_hist = json.dumps(
            [m['total_litros'] for m in hist_mensual] + [None] * 6
        )
        hist_proy_proy = json.dumps(
            [None] * len(hist_mensual) + proy_valores
        )
        # Punto de unión entre histórico y proyección
        if hist_mensual:
            hist_proy_union = json.dumps(
                [None] * (len(hist_mensual) - 1) + [hist_mensual[-1]['total_litros']] + proy_valores[:1]
            )
        else:
            hist_proy_union = json.dumps([None])

        context = {
            **self.admin_site.each_context(request),
            'title': 'Estadísticas de Combustible',
            'opts': self.model._meta,
            'has_permission': True,
            # Filtros
            'desde': desde.isoformat(),
            'hasta': hasta.isoformat(),
            # KPIs
            'totales': totales,
            'estados_count': estados_count,
            'total_alertas_pendientes': total_alertas_pendientes,
            # Nivel inicial
            'niveles': niveles,
            'nivel_labels_json': nivel_labels_json,
            'nivel_totales_json': nivel_totales_json,
            'nivel_promedios_json': nivel_promedios_json,
            'nivel_litros_json': nivel_litros_json,
            # Por día
            'dias_labels_json': dias_labels_json,
            'dias_litros_json': dias_litros_json,
            'dias_cargas_json': dias_cargas_json,
            # Top unidades
            'top_unidades': top_unidades,
            'top_unidades_labels_json': top_unidades_labels_json,
            'top_unidades_litros_json': top_unidades_litros_json,
            'top_unidades_cargas_json': top_unidades_cargas_json,
            # Candados
            'candados': candados,
            'candados_labels_json': candados_labels_json,
            'candados_totales_json': candados_totales_json,
            # Despachadores
            'por_desp': por_desp,
            'desp_labels_json': desp_labels_json,
            'desp_litros_json': desp_litros_json,
            'desp_cargas_json': desp_cargas_json,
            # Historial mensual
            'hist_mensual': hist_mensual,
            'hist_labels_json': hist_labels_json,
            'hist_litros_json': hist_litros_json,
            'hist_cargas_json': hist_cargas_json,
            # Alertas
            'alertas': alertas,
            'alertas_labels_json': alertas_labels_json,
            'alertas_totales_json': alertas_totales_json,
            'alertas_pendientes_json': alertas_pendientes_json,
            # Proyecciones
            'proyeccion_30_dias': proyeccion_30_dias,
            'proyeccion_90_dias': proyeccion_90_dias,
            'proyeccion_anual': proyeccion_anual,
            'proyeccion_cargas_30': int(proyeccion_cargas_30),
            'promedio_diario_litros': round(promedio_diario_litros, 2),
            'variacion_vs_anterior': variacion_vs_anterior,
            'litros_periodo_anterior': round(litros_periodo_anterior, 2),
            'tendencia_label': tendencia_label,
            'tendencia_color': tendencia_color,
            'variacion_interna': variacion_interna,
            'hist_proy_labels': hist_proy_labels,
            'hist_proy_hist': hist_proy_hist,
            'hist_proy_proy': hist_proy_proy,
            'hist_proy_union': hist_proy_union,
            # Interpretaciones
            'interpretaciones': interpretaciones,
        }

        return render(request, 'admin/combustible/reporte_estadisticas.html', context)

    # ─── Exportar Excel ────────────────────────────────────────────────────────

    def exportar_excel_view(self, request):
        hoy = date.today()
        desde_str = request.GET.get('desde', (hoy - timedelta(days=30)).isoformat())
        hasta_str = request.GET.get('hasta', hoy.isoformat())
        try:
            desde = date.fromisoformat(desde_str)
        except (ValueError, TypeError):
            desde = hoy - timedelta(days=30)
        try:
            hasta = date.fromisoformat(hasta_str)
        except (ValueError, TypeError):
            hasta = hoy

        qs = CargaCombustible.objects.filter(
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=desde,
            fecha_hora_inicio__date__lte=hasta,
        ).select_related('unidad', 'despachador')

        wb = openpyxl.Workbook()

        # ── Estilos reutilizables ──────────────────────────────────────────────
        color_azul = '1F4E79'
        color_verde = '1A6B3C'
        color_fila_alt = 'EBF3FB'
        color_alerta = 'FFF3CD'
        color_peligro = 'FDECEA'

        def estilo_header(ws, row, cols, color=color_azul):
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            font = Font(bold=True, color='FFFFFF', size=11)
            aln = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin = Side(border_style='thin', color='FFFFFF')
            brd = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = aln
                cell.border = brd
            ws.row_dimensions[row].height = 30

        def estilo_fila(ws, row, cols, alt=False, color_fill=None):
            fill_color = color_fill or (color_fila_alt if alt else 'FFFFFF')
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            thin = Side(border_style='thin', color='D0D0D0')
            brd = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = brd
                cell.alignment = Alignment(vertical='center')

        def titulo_hoja(ws, texto, cols):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
            cell = ws['A1']
            cell.value = texto
            cell.font = Font(bold=True, color='FFFFFF', size=14)
            cell.fill = PatternFill(start_color=color_azul, end_color=color_azul, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
            ws['A2'].value = f'Período: {desde.strftime("%d/%m/%Y")} — {hasta.strftime("%d/%m/%Y")}'
            ws['A2'].font = Font(italic=True, color='555555', size=10)
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 20

        # ── Hoja 1: Resumen general ────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Resumen'
        titulo_hoja(ws1, '⛽ Reporte de Combustible — Resumen General', 4)

        totales = qs.aggregate(
            total_cargas=Count('id'),
            total_litros=Sum('cantidad_litros'),
            promedio_litros=Avg('cantidad_litros'),
            promedio_tiempo=Avg('tiempo_carga_minutos'),
            max_litros=Max('cantidad_litros'),
            min_litros=Min('cantidad_litros'),
        )

        kpis = [
            ('Total cargas completadas', totales['total_cargas'] or 0),
            ('Total litros cargados', round(float(totales['total_litros'] or 0), 2)),
            ('Promedio litros por carga', round(float(totales['promedio_litros'] or 0), 2)),
            ('Carga máxima (litros)', round(float(totales['max_litros'] or 0), 2)),
            ('Carga mínima (litros)', round(float(totales['min_litros'] or 0), 2)),
            ('Tiempo promedio de carga (min)', round(float(totales['promedio_tiempo'] or 0), 1)),
            ('Alertas pendientes en el período', AlertaCombustible.objects.filter(
                resuelta=False,
                carga__fecha_hora_inicio__date__gte=desde,
                carga__fecha_hora_inicio__date__lte=hasta,
            ).count()),
        ]

        ws1['A4'] = 'Indicador'
        ws1['B4'] = 'Valor'
        estilo_header(ws1, 4, 2)
        for i, (indicador, valor) in enumerate(kpis, start=5):
            ws1['A' + str(i)] = indicador
            ws1['B' + str(i)] = valor
            estilo_fila(ws1, i, 2, alt=(i % 2 == 0))
            ws1['A' + str(i)].font = Font(bold=False, size=11)
            ws1['B' + str(i)].alignment = Alignment(horizontal='center')

        ws1.column_dimensions['A'].width = 40
        ws1.column_dimensions['B'].width = 20

        # ─ Niveles iniciales en resumen
        NIVEL_LABELS = {'VACIO': 'Vacío', 'CUARTO': '1/4', 'MEDIO': '1/2', 'TRES_CUARTOS': '3/4'}
        fila_base = 5 + len(kpis) + 2

        ws1.cell(row=fila_base, column=1).value = 'Distribución por Nivel Inicial'
        ws1.cell(row=fila_base, column=2).value = '# Cargas'
        ws1.cell(row=fila_base, column=3).value = 'Total litros'
        ws1.cell(row=fila_base, column=4).value = 'Promedio litros'
        estilo_header(ws1, fila_base, 4)
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 18

        niveles_raw = qs.values('nivel_combustible_inicial').annotate(
            total=Count('id'), litros=Sum('cantidad_litros'), promedio=Avg('cantidad_litros')
        )
        for i, n in enumerate(niveles_raw, start=fila_base + 1):
            ws1.cell(row=i, column=1).value = NIVEL_LABELS.get(n['nivel_combustible_inicial'], n['nivel_combustible_inicial'])
            ws1.cell(row=i, column=2).value = n['total']
            ws1.cell(row=i, column=3).value = round(float(n['litros'] or 0), 2)
            ws1.cell(row=i, column=4).value = round(float(n['promedio'] or 0), 2)
            estilo_fila(ws1, i, 4, alt=(i % 2 == 0))

        # ── Hoja 2: Cargas detalle ─────────────────────────────────────────────
        ws2 = wb.create_sheet('Cargas Detalle')
        COLS2 = 14
        titulo_hoja(ws2, '⛽ Detalle de Cargas de Combustible', COLS2)

        headers2 = [
            'ID', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Unidad', 'Placa',
            'Tipo Unidad', 'Despachador', 'Litros', 'Kilometraje',
            'Nivel Inicial', 'Estado Candado', 'Tiempo (min)', 'Notas',
        ]
        for col, h in enumerate(headers2, start=1):
            ws2.cell(row=3, column=col).value = h
        estilo_header(ws2, 3, COLS2)

        anchos2 = [6, 12, 12, 12, 12, 12, 12, 18, 10, 14, 14, 14, 12, 30]
        for col, ancho in enumerate(anchos2, start=1):
            ws2.column_dimensions[get_column_letter(col)].width = ancho

        CANDADO_LABELS = {'NORMAL': 'Normal', 'ALTERADO': 'Alterado', 'VIOLADO': 'Violado', 'SIN_CANDADO': 'Sin Candado'}
        for i, carga in enumerate(qs.order_by('-fecha_hora_inicio'), start=4):
            data = [
                carga.id,
                carga.fecha_hora_inicio.strftime('%d/%m/%Y'),
                carga.fecha_hora_inicio.strftime('%H:%M'),
                carga.fecha_hora_fin.strftime('%H:%M') if carga.fecha_hora_fin else '',
                carga.unidad.numero_economico,
                carga.unidad.placa,
                carga.unidad.get_tipo_display(),
                carga.despachador.nombre,
                float(carga.cantidad_litros),
                carga.kilometraje_actual,
                NIVEL_LABELS.get(carga.nivel_combustible_inicial, carga.nivel_combustible_inicial),
                CANDADO_LABELS.get(carga.estado_candado_anterior, carga.estado_candado_anterior),
                carga.tiempo_carga_minutos or '',
                carga.notas,
            ]
            for col, val in enumerate(data, start=1):
                ws2.cell(row=i, column=col).value = val

            fill_color = color_alerta if carga.tiene_alertas() else None
            estilo_fila(ws2, i, COLS2, alt=(i % 2 == 0), color_fill=fill_color)

        # ── Hoja 3: Por unidad ─────────────────────────────────────────────────
        ws3 = wb.create_sheet('Por Unidad')
        COLS3 = 7
        titulo_hoja(ws3, '🚛 Estadísticas por Unidad', COLS3)

        headers3 = ['N° Económico', 'Placa', 'Tipo', '# Cargas', 'Total Litros', 'Promedio Litros', 'Máx. Litros']
        for col, h in enumerate(headers3, start=1):
            ws3.cell(row=3, column=col).value = h
        estilo_header(ws3, 3, COLS3)

        anchos3 = [16, 14, 14, 12, 16, 16, 14]
        for col, ancho in enumerate(anchos3, start=1):
            ws3.column_dimensions[get_column_letter(col)].width = ancho

        por_unidad = (
            qs.values('unidad__numero_economico', 'unidad__placa', 'unidad__tipo')
            .annotate(
                total_cargas=Count('id'),
                total_litros=Sum('cantidad_litros'),
                promedio_litros=Avg('cantidad_litros'),
                max_litros=Max('cantidad_litros'),
            )
            .order_by('-total_litros')
        )
        for i, u in enumerate(por_unidad, start=4):
            row_data = [
                u['unidad__numero_economico'],
                u['unidad__placa'],
                u['unidad__tipo'],
                u['total_cargas'],
                round(float(u['total_litros'] or 0), 2),
                round(float(u['promedio_litros'] or 0), 2),
                round(float(u['max_litros'] or 0), 2),
            ]
            for col, val in enumerate(row_data, start=1):
                ws3.cell(row=i, column=col).value = val
            estilo_fila(ws3, i, COLS3, alt=(i % 2 == 0))

        # ── Hoja 4: Por despachador ────────────────────────────────────────────
        ws4 = wb.create_sheet('Por Despachador')
        COLS4 = 5
        titulo_hoja(ws4, '👤 Estadísticas por Despachador', COLS4)

        headers4 = ['Despachador', '# Cargas', 'Total Litros', 'Promedio Litros', 'Tiempo Prom. (min)']
        for col, h in enumerate(headers4, start=1):
            ws4.cell(row=3, column=col).value = h
        estilo_header(ws4, 3, COLS4, color=color_verde)

        anchos4 = [24, 12, 16, 16, 18]
        for col, ancho in enumerate(anchos4, start=1):
            ws4.column_dimensions[get_column_letter(col)].width = ancho

        por_desp = (
            qs.values('despachador__nombre')
            .annotate(
                total_cargas=Count('id'),
                total_litros=Sum('cantidad_litros'),
                promedio_litros=Avg('cantidad_litros'),
                promedio_tiempo=Avg('tiempo_carga_minutos'),
            )
            .order_by('-total_litros')
        )
        for i, d in enumerate(por_desp, start=4):
            row_data = [
                d['despachador__nombre'],
                d['total_cargas'],
                round(float(d['total_litros'] or 0), 2),
                round(float(d['promedio_litros'] or 0), 2),
                round(float(d['promedio_tiempo'] or 0), 1),
            ]
            for col, val in enumerate(row_data, start=1):
                ws4.cell(row=i, column=col).value = val
            estilo_fila(ws4, i, COLS4, alt=(i % 2 == 0))

        # ── Hoja 5: Alertas ────────────────────────────────────────────────────
        ws5 = wb.create_sheet('Alertas')
        COLS5 = 7
        titulo_hoja(ws5, '⚠️ Alertas de Combustible', COLS5)

        headers5 = ['Fecha', 'Unidad', 'Tipo de Alerta', 'Mensaje', 'Resuelta', 'Resuelta por', 'Fecha Resolución']
        for col, h in enumerate(headers5, start=1):
            ws5.cell(row=3, column=col).value = h
        estilo_header(ws5, 3, COLS5, color='B45309')

        anchos5 = [14, 14, 22, 40, 10, 20, 16]
        for col, ancho in enumerate(anchos5, start=1):
            ws5.column_dimensions[get_column_letter(col)].width = ancho

        ALERTA_LABELS = dict(AlertaCombustible.TIPO_CHOICES)
        alertas_qs = AlertaCombustible.objects.filter(
            carga__fecha_hora_inicio__date__gte=desde,
            carga__fecha_hora_inicio__date__lte=hasta,
        ).select_related('carga__unidad', 'resuelta_por').order_by('-fecha_generacion')

        for i, alerta in enumerate(alertas_qs, start=4):
            row_data = [
                alerta.fecha_generacion.strftime('%d/%m/%Y %H:%M'),
                alerta.carga.unidad.numero_economico,
                ALERTA_LABELS.get(alerta.tipo_alerta, alerta.tipo_alerta),
                alerta.mensaje,
                'Sí' if alerta.resuelta else 'No',
                alerta.resuelta_por.get_full_name() if alerta.resuelta_por else '',
                alerta.fecha_resolucion.strftime('%d/%m/%Y') if alerta.fecha_resolucion else '',
            ]
            for col, val in enumerate(row_data, start=1):
                ws5.cell(row=i, column=col).value = val
            fill_color = None if alerta.resuelta else color_peligro
            estilo_fila(ws5, i, COLS5, alt=(i % 2 == 0), color_fill=fill_color)

        # ── Respuesta HTTP ─────────────────────────────────────────────────────
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="reporte_combustible_{desde.strftime("%Y%m%d")}_{hasta.strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response
