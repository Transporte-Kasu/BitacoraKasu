from datetime import date, timedelta

from django.contrib import admin
from django.utils.html import format_html
from django.utils import timezone
from django.shortcuts import render
from django.contrib import messages
from django.urls import path
from django.db.models import Sum
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .models import Unidad


def _suma_cantidad_por_costo(queryset):
    """Suma cantidad * producto.costo_unitario sobre un queryset de salidas de almacén."""
    total = Decimal('0')
    for item in queryset.select_related('producto'):
        total += item.cantidad * item.producto.costo_unitario
    return total


def calcular_reporte_utilidad(desde, hasta):
    """
    Calcula ingreso, gasto y utilidad por unidad activa en el rango [desde, hasta].

    Gasto = combustible (CargaCombustible.costo_calculado) + taller (OrdenTrabajo.costo_total_real)
    + consumibles de almacén (SalidaRapidaConsumible + AsignacionDirectaAlmacen +
    ItemAsignacionSalida con tipo_destino='UNIDAD'). Las piezas de taller vía SalidaAlmacen
    ligada a OrdenTrabajo no se incluyen aquí para no duplicar el gasto ya contado en
    OrdenTrabajo.costo_total_real.
    """
    from modulos.almacen.models import (
        SalidaRapidaConsumible, AsignacionDirectaAlmacen, ItemAsignacionSalida,
    )
    from modulos.bitacoras.models import BitacoraViaje
    from modulos.combustible.models import CargaCombustible
    from modulos.taller.models import OrdenTrabajo

    filas = []
    for unidad in Unidad.objects.filter(activa=True).order_by('numero_economico'):
        ingresos = BitacoraViaje.objects.filter(
            unidad=unidad, completado=True,
            fecha_llegada__date__gte=desde, fecha_llegada__date__lte=hasta,
        ).aggregate(t=Sum('ingreso_calculado'))['t'] or Decimal('0')

        gasto_combustible = CargaCombustible.objects.filter(
            unidad=unidad, estado='COMPLETADO',
            fecha_hora_inicio__date__gte=desde, fecha_hora_inicio__date__lte=hasta,
        ).aggregate(t=Sum('costo_calculado'))['t'] or Decimal('0')

        ordenes_completadas = OrdenTrabajo.objects.filter(
            unidad=unidad, estado='COMPLETADA',
            fecha_finalizacion__date__gte=desde, fecha_finalizacion__date__lte=hasta,
        )
        gasto_taller = sum((orden.costo_total_real for orden in ordenes_completadas), Decimal('0'))

        gasto_consumibles = (
            _suma_cantidad_por_costo(SalidaRapidaConsumible.objects.filter(
                unidad=unidad, fecha_salida__date__gte=desde, fecha_salida__date__lte=hasta,
            ))
            + _suma_cantidad_por_costo(AsignacionDirectaAlmacen.objects.filter(
                unidad=unidad, fecha_asignacion__date__gte=desde, fecha_asignacion__date__lte=hasta,
            ))
            + _suma_cantidad_por_costo(ItemAsignacionSalida.objects.filter(
                asignacion__tipo_destino='UNIDAD', asignacion__unidad=unidad,
                asignacion__fecha__gte=desde, asignacion__fecha__lte=hasta,
            ))
        )

        gasto_total = gasto_combustible + gasto_taller + gasto_consumibles
        utilidad = ingresos - gasto_total
        utilidad_pct = (utilidad / ingresos * 100) if ingresos else None

        filas.append({
            'unidad': unidad,
            'ingresos': ingresos,
            'gasto_combustible': gasto_combustible,
            'gasto_taller': gasto_taller,
            'gasto_consumibles': gasto_consumibles,
            'gasto_total': gasto_total,
            'utilidad': utilidad,
            'utilidad_pct': utilidad_pct,
        })

    totales = {
        clave: sum((f[clave] for f in filas), Decimal('0'))
        for clave in ('ingresos', 'gasto_combustible', 'gasto_taller', 'gasto_consumibles', 'gasto_total', 'utilidad')
    }

    bitacoras_excluidas = BitacoraViaje.objects.filter(
        completado=True, fecha_llegada__date__gte=desde, fecha_llegada__date__lte=hasta,
        ingreso_calculado__isnull=True,
    ).count()
    cargas_excluidas = CargaCombustible.objects.filter(
        estado='COMPLETADO', fecha_hora_inicio__date__gte=desde, fecha_hora_inicio__date__lte=hasta,
        costo_calculado__isnull=True,
    ).count()

    return {
        'filas': filas,
        'totales': totales,
        'bitacoras_excluidas': bitacoras_excluidas,
        'cargas_excluidas': cargas_excluidas,
    }


@admin.register(Unidad)
class UnidadAdmin(admin.ModelAdmin):
    change_list_template = 'admin/unidades/unidad/change_list.html'
    list_display = [
        'numero_economico', 'placa', 'tipo_badge', 'marca', 'modelo', 'año',
        'kilometraje_actual', 'activa_badge', 'mantenimiento_status',
        'viajes_completados_display'
    ]
    list_filter = ['tipo', 'activa', 'marca', 'año']
    search_fields = ['numero_economico', 'placa', 'marca', 'modelo']
    readonly_fields = [
        'fecha_alta', 'created_at', 'updated_at',
        'rendimiento_promedio_real_display', 'eficiencia_combustible_display'
    ]
    
    fieldsets = (
        ('Identificación', {
            'fields': ('numero_economico', 'placa', 'tipo')
        }),
        ('Especificaciones Técnicas', {
            'fields': ('marca', 'modelo', 'año', 'capacidad_combustible', 'rendimiento_esperado')
        }),
        ('Kilometraje', {
            'fields': (
                'kilometraje_actual',
                'rendimiento_promedio_real_display',
                'eficiencia_combustible_display'
            )
        }),
        ('Mantenimiento', {
            'fields': ('ultimo_mantenimiento', 'proximo_mantenimiento')
        }),
        ('Estado', {
            'fields': ('activa', 'fecha_alta', 'fecha_baja')
        }),
        ('Control de Combustible', {
            'fields': ('control_combustible_total',),
            'description': 'Configuración futura: forzar proceso completo en unidades locales.'
        }),
        ('Notas', {
            'fields': ('notas',)
        }),
        ('Metadatos', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def tipo_badge(self, obj):
        """Badge con el tipo de unidad"""
        colors = {
            'LOCAL': '#3b82f6',
            'FORANEA': '#8b5cf6',
            'ESPERANZA': '#10b981',
        }
        color = colors.get(obj.tipo, '#6b7280')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; '
            'border-radius: 3px; font-weight: bold;">{}</span>',
            color, obj.get_tipo_display()
        )
    tipo_badge.short_description = 'Tipo'
    
    def activa_badge(self, obj):
        """Badge de estado activo/inactivo"""
        if obj.activa:
            return format_html(
                '<span style="background-color: #10b981; color: white; padding: 3px 10px; '
                'border-radius: 3px; font-weight: bold;">✓ Activa</span>'
            )
        else:
            return format_html(
                '<span style="background-color: #ef4444; color: white; padding: 3px 10px; '
                'border-radius: 3px; font-weight: bold;">✗ Inactiva</span>'
            )
    activa_badge.short_description = 'Estado'
    
    def mantenimiento_status(self, obj):
        """Muestra el estado del mantenimiento"""
        if obj.requiere_mantenimiento():
            return format_html(
                '<span style="background-color: #ef4444; color: white; padding: 3px 10px; '
                'border-radius: 3px; font-weight: bold;">⚠ Requiere</span>'
            )
        elif obj.proximo_mantenimiento:
            dias_restantes = (obj.proximo_mantenimiento - timezone.now().date()).days
            if dias_restantes <= 7:
                color = '#fbbf24'  # Amarillo
                texto = f'{dias_restantes} días'
            else:
                color = '#10b981'  # Verde
                texto = f'{dias_restantes} días'
            return format_html(
                '<span style="background-color: {}; color: white; padding: 3px 10px; '
                'border-radius: 3px; font-weight: bold;">{}</span>',
                color, texto
            )
        return format_html('<span style="color: #6b7280;">Sin programar</span>')
    mantenimiento_status.short_description = 'Mantenimiento'
    
    def viajes_completados_display(self, obj):
        """Muestra el número de viajes completados"""
        count = obj.viajes_completados()
        return format_html(
            '<span style="background: #3b82f6; color: white; padding: 2px 8px; '
            'border-radius: 4px; font-weight: bold;">{}</span>',
            count
        )
    viajes_completados_display.short_description = 'Viajes'
    
    def rendimiento_promedio_real_display(self, obj):
        """Muestra el rendimiento promedio real"""
        rendimiento = obj.rendimiento_promedio_real()
        if rendimiento > 0:
            return f"{rendimiento} km/lt"
        return "Sin datos"
    rendimiento_promedio_real_display.short_description = 'Rendimiento Real'
    
    def eficiencia_combustible_display(self, obj):
        """Muestra la eficiencia de combustible"""
        eficiencia = obj.eficiencia_combustible()
        if eficiencia > 0:
            if eficiencia >= 90:
                color = '#10b981'
            elif eficiencia >= 70:
                color = '#fbbf24'
            else:
                color = '#ef4444'
            return format_html(
                '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
                color, eficiencia
            )
        return "Sin datos"
    eficiencia_combustible_display.short_description = 'Eficiencia'
    
    actions = ['activar_unidades', 'desactivar_unidades', 'actualizar_rendimiento', 'actualizar_capacidad_tanque']
    
    def activar_unidades(self, request, queryset):
        """Activa las unidades seleccionadas"""
        updated = queryset.update(activa=True, fecha_baja=None)
        self.message_user(request, f'{updated} unidades activadas.')
    activar_unidades.short_description = 'Activar unidades seleccionadas'
    
    def desactivar_unidades(self, request, queryset):
        """Desactiva las unidades seleccionadas"""
        updated = queryset.update(activa=False, fecha_baja=timezone.now().date())
        self.message_user(request, f'{updated} unidades desactivadas.')
    desactivar_unidades.short_description = 'Desactivar unidades seleccionadas'

    def actualizar_rendimiento(self, request, queryset):
        """Actualizar rendimiento esperado de combustible a varias unidades"""
        if 'confirmado' in request.POST:
            try:
                nuevo_rendimiento = Decimal(request.POST.get('rendimiento_esperado', '0'))
            except (InvalidOperation, ValueError):
                self.message_user(
                    request, 'Valor de rendimiento inválido.', messages.ERROR
                )
                return

            if nuevo_rendimiento <= 0:
                self.message_user(
                    request, 'El rendimiento debe ser mayor a 0.', messages.ERROR
                )
                return

            updated = queryset.update(rendimiento_esperado=nuevo_rendimiento)
            self.message_user(
                request,
                f'Rendimiento actualizado a {nuevo_rendimiento} km/lt '
                f'en {updated} unidad(es).',
                messages.SUCCESS
            )
            return

        return render(
            request,
            'admin/unidades/actualizar_rendimiento.html',
            {
                'unidades': queryset.order_by('numero_economico'),
                'opts': self.model._meta,
                'title': 'Actualizar Rendimiento de Combustible',
            }
        )
    actualizar_rendimiento.short_description = '⛽ Actualizar rendimiento de combustible'

    def actualizar_capacidad_tanque(self, request, queryset):
        """Establece la capacidad del tanque de combustible a varias unidades"""
        if 'confirmado' in request.POST:
            try:
                nueva_capacidad = Decimal(request.POST.get('capacidad_combustible', '0'))
            except (InvalidOperation, ValueError):
                self.message_user(
                    request, 'Valor de capacidad inválido.', messages.ERROR
                )
                return

            if nueva_capacidad <= 0:
                self.message_user(
                    request, 'La capacidad debe ser mayor a 0.', messages.ERROR
                )
                return

            updated = queryset.update(capacidad_combustible=nueva_capacidad)
            self.message_user(
                request,
                f'Capacidad de tanque actualizada a {nueva_capacidad} lt en {updated} unidad(es).',
                messages.SUCCESS
            )
            return

        return render(
            request,
            'admin/unidades/actualizar_capacidad_tanque.html',
            {
                'unidades': queryset.order_by('numero_economico'),
                'opts': self.model._meta,
                'title': 'Actualizar Capacidad de Tanque',
            }
        )
    actualizar_capacidad_tanque.short_description = '🛢️ Actualizar capacidad de tanque'

    # ─── Reporte de utilidad por unidad ────────────────────────────────────────

    def _rango_fechas(self, request):
        hoy = date.today()
        primer_dia_mes = hoy.replace(day=1)
        desde_str = request.GET.get('desde', primer_dia_mes.isoformat())
        hasta_str = request.GET.get('hasta', hoy.isoformat())
        try:
            desde = date.fromisoformat(desde_str)
        except (ValueError, TypeError):
            desde = primer_dia_mes
        try:
            hasta = date.fromisoformat(hasta_str)
        except (ValueError, TypeError):
            hasta = hoy
        return desde, hasta

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'reporte-utilidad/',
                self.admin_site.admin_view(self.reporte_utilidad_view),
                name='unidades_unidad_reporte_utilidad',
            ),
            path(
                'reporte-utilidad/exportar-excel/',
                self.admin_site.admin_view(self.exportar_excel_reporte_utilidad_view),
                name='unidades_unidad_reporte_utilidad_excel',
            ),
        ]
        return custom_urls + urls

    def reporte_utilidad_view(self, request):
        desde, hasta = self._rango_fechas(request)
        resultado = calcular_reporte_utilidad(desde, hasta)

        return render(
            request,
            'admin/unidades/reporte_utilidad.html',
            {
                **self.admin_site.each_context(request),
                'title': 'Reporte de Utilidad por Unidad',
                'opts': self.model._meta,
                'filas': resultado['filas'],
                'totales': resultado['totales'],
                'bitacoras_excluidas': resultado['bitacoras_excluidas'],
                'cargas_excluidas': resultado['cargas_excluidas'],
                'desde': desde,
                'hasta': hasta,
            }
        )

    def exportar_excel_reporte_utilidad_view(self, request):
        desde, hasta = self._rango_fechas(request)
        resultado = calcular_reporte_utilidad(desde, hasta)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Utilidad por Unidad'

        color_azul = '1F4E79'
        fill_header = PatternFill(start_color=color_azul, end_color=color_azul, fill_type='solid')
        font_header = Font(bold=True, color='FFFFFF')
        thin = Side(border_style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:H1')
        ws['A1'] = f'Reporte de Utilidad por Unidad — {desde.isoformat()} a {hasta.isoformat()}'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = fill_header
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        encabezados = [
            'Unidad', 'Ingresos', 'Gasto Combustible', 'Gasto Taller',
            'Gasto Consumibles', 'Gasto Total', 'Utilidad $', 'Utilidad %',
        ]
        for col, texto in enumerate(encabezados, start=1):
            cell = ws.cell(row=2, column=col, value=texto)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        fila_actual = 3
        for fila in resultado['filas']:
            valores = [
                fila['unidad'].numero_economico,
                float(fila['ingresos']),
                float(fila['gasto_combustible']),
                float(fila['gasto_taller']),
                float(fila['gasto_consumibles']),
                float(fila['gasto_total']),
                float(fila['utilidad']),
                round(float(fila['utilidad_pct']), 2) if fila['utilidad_pct'] is not None else None,
            ]
            for col, valor in enumerate(valores, start=1):
                cell = ws.cell(row=fila_actual, column=col, value=valor)
                cell.border = border
            fila_actual += 1

        totales = resultado['totales']
        valores_totales = [
            'TOTAL',
            float(totales['ingresos']),
            float(totales['gasto_combustible']),
            float(totales['gasto_taller']),
            float(totales['gasto_consumibles']),
            float(totales['gasto_total']),
            float(totales['utilidad']),
            None,
        ]
        for col, valor in enumerate(valores_totales, start=1):
            cell = ws.cell(row=fila_actual, column=col, value=valor)
            cell.font = Font(bold=True)
            cell.border = border

        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="utilidad_por_unidad_{desde}_{hasta}.xlsx"'
        wb.save(response)
        return response
